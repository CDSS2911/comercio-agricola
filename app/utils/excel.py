from datetime import date, datetime
from io import BytesIO

from flask import Response
from openpyxl import Workbook
from openpyxl.styles import Font


def _normalize_excel_value(value):
    if isinstance(value, (datetime, date)):
        return value
    if value is None:
        return ""
    return value


def create_excel_response(filename, sheet_name, headers, rows):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]

    ws.append(headers)
    for col in ws[1]:
        col.font = Font(bold=True)

    for row in rows:
        ws.append([_normalize_excel_value(v) for v in row])

    for column_cells in ws.columns:
        max_length = 0
        col_letter = column_cells[0].column_letter
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        ws.column_dimensions[col_letter].width = min(max(12, max_length + 2), 45)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = Response(
        output.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response.headers["Content-Disposition"] = f"attachment; filename={filename}"
    return response


def create_excel_multisheet_response(filename, sheets):
    """
    sheets: list of dicts with keys:
      - name
      - headers
      - rows
    """
    wb = Workbook()
    first = True

    for sheet in sheets:
        if first:
            ws = wb.active
            ws.title = sheet["name"][:31]
            first = False
        else:
            ws = wb.create_sheet(title=sheet["name"][:31])

        headers = sheet.get("headers", [])
        rows = sheet.get("rows", [])

        ws.append(headers)
        for col in ws[1]:
            col.font = Font(bold=True)

        for row in rows:
            ws.append([_normalize_excel_value(v) for v in row])

        for column_cells in ws.columns:
            max_length = 0
            col_letter = column_cells[0].column_letter
            for cell in column_cells:
                value = "" if cell.value is None else str(cell.value)
                max_length = max(max_length, len(value))
            ws.column_dimensions[col_letter].width = min(max(12, max_length + 2), 45)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = Response(
        output.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response.headers["Content-Disposition"] = f"attachment; filename={filename}"
    return response
