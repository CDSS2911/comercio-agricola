from flask import render_template, redirect, url_for, flash, request, jsonify, make_response
from flask_login import login_required, current_user
from sqlalchemy import func, extract
from datetime import datetime, timedelta, date
import base64
import csv
import io
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from PIL import Image as PILImage
from app.routes import bp
from app.models import (
    User,
    LoteRecoleccion,
    Huevo,
    CategoriaHuevo,
    Venta,
    DetalleVenta,
    RegistroMortalidad,
    VentaGallinas,
)
from app import db


@bp.route('/')
@bp.route('/index')
def index():
    """Página principal"""
    return render_template('index.html', title='Inicio')


@bp.route('/dashboard')
@login_required
def dashboard():
    """Panel de control del usuario"""
    # Obtener datos de inventario para gráficos
    hoy = datetime.now().date()
    hace_30_dias = hoy - timedelta(days=30)
    
    # Datos diarios de los últimos 30 días - usar consultas separadas por simplicidad
    datos_diarios_raw = db.session.query(
        func.date(LoteRecoleccion.fecha_recoleccion).label('fecha'),
        func.count(Huevo.id).label('total_huevos')
    ).join(
        Huevo, LoteRecoleccion.id == Huevo.lote_id
    ).filter(
        LoteRecoleccion.fecha_recoleccion >= hace_30_dias
    ).group_by(
        func.date(LoteRecoleccion.fecha_recoleccion)
    ).order_by(
        func.date(LoteRecoleccion.fecha_recoleccion).desc()
    ).all()
    
    # Crear estructura de datos más simple
    datos_diarios = []
    for dato in datos_diarios_raw:
        # Obtener conteos específicos por fecha
        fecha = dato.fecha
        huevos_buenos = db.session.query(func.count(Huevo.id)).join(LoteRecoleccion).filter(
            func.date(LoteRecoleccion.fecha_recoleccion) == fecha,
            Huevo.roto == False
        ).scalar() or 0
        
        huevos_rotos = db.session.query(func.count(Huevo.id)).join(LoteRecoleccion).filter(
            func.date(LoteRecoleccion.fecha_recoleccion) == fecha,
            Huevo.roto == True
        ).scalar() or 0
        
        peso_total = db.session.query(func.sum(Huevo.peso)).join(LoteRecoleccion).filter(
            func.date(LoteRecoleccion.fecha_recoleccion) == fecha,
            Huevo.roto == False
        ).scalar() or 0
        
        # Crear objeto con estructura similar
        class DatoDiario:
            def __init__(self, fecha, total, buenos, rotos, peso):
                self.fecha = fecha
                self.total_huevos = total
                self.huevos_buenos = buenos
                self.huevos_rotos = rotos
                self.peso_total = peso
                
        datos_diarios.append(DatoDiario(fecha, dato.total_huevos, huevos_buenos, huevos_rotos, peso_total))
    
    # Datos mensuales del año actual - usar consultas separadas por simplicidad
    año_actual = hoy.year
    datos_mensuales_raw = db.session.query(
        extract('month', LoteRecoleccion.fecha_recoleccion).label('mes'),
        func.count(Huevo.id).label('total_huevos')
    ).join(
        Huevo, LoteRecoleccion.id == Huevo.lote_id
    ).filter(
        extract('year', LoteRecoleccion.fecha_recoleccion) == año_actual
    ).group_by(
        extract('month', LoteRecoleccion.fecha_recoleccion)
    ).order_by(
        extract('month', LoteRecoleccion.fecha_recoleccion)
    ).all()
    
    # Crear estructura de datos más simple
    datos_mensuales = []
    for dato in datos_mensuales_raw:
        # Obtener conteos específicos por mes
        mes = dato.mes
        huevos_buenos = db.session.query(func.count(Huevo.id)).join(LoteRecoleccion).filter(
            extract('year', LoteRecoleccion.fecha_recoleccion) == año_actual,
            extract('month', LoteRecoleccion.fecha_recoleccion) == mes,
            Huevo.roto == False
        ).scalar() or 0
        
        huevos_rotos = db.session.query(func.count(Huevo.id)).join(LoteRecoleccion).filter(
            extract('year', LoteRecoleccion.fecha_recoleccion) == año_actual,
            extract('month', LoteRecoleccion.fecha_recoleccion) == mes,
            Huevo.roto == True
        ).scalar() or 0
        
        peso_total = db.session.query(func.sum(Huevo.peso)).join(LoteRecoleccion).filter(
            extract('year', LoteRecoleccion.fecha_recoleccion) == año_actual,
            extract('month', LoteRecoleccion.fecha_recoleccion) == mes,
            Huevo.roto == False
        ).scalar() or 0
        
        # Crear objeto con estructura similar
        class DatoMensual:
            def __init__(self, mes, total, buenos, rotos, peso):
                self.mes = mes
                self.total_huevos = total
                self.huevos_buenos = buenos
                self.huevos_rotos = rotos
                self.peso_total = peso
                
        datos_mensuales.append(DatoMensual(mes, dato.total_huevos, huevos_buenos, huevos_rotos, peso_total))
    
    # Distribución por categorías
    distribucion_categorias = db.session.query(
        CategoriaHuevo.nombre,
        CategoriaHuevo.peso_min,
        CategoriaHuevo.peso_max,
        func.count(Huevo.id).label('cantidad'),
        func.sum(Huevo.peso).label('peso_total')
    ).join(
        Huevo, CategoriaHuevo.id == Huevo.categoria_id
    ).filter(
        Huevo.roto == False,
        Huevo.vendido == False
    ).group_by(
        CategoriaHuevo.id, CategoriaHuevo.nombre, CategoriaHuevo.peso_min, CategoriaHuevo.peso_max
    ).all()
    
    # Estadísticas generales
    stats_generales = {
        'total_lotes': LoteRecoleccion.query.count(),
        'total_huevos': Huevo.query.filter(Huevo.roto == False).count(),
        'huevos_rotos': Huevo.query.filter(Huevo.roto == True).count(),
        'huevos_disponibles': Huevo.query.filter(Huevo.roto == False, Huevo.vendido == False).count(),
        'peso_total': db.session.query(func.sum(Huevo.peso)).filter(Huevo.roto == False).scalar() or 0
    }

    meses_nombres = ['Ene', 'Feb', 'Mar', 'Abr', 'May', 'Jun',
                     'Jul', 'Ago', 'Sep', 'Oct', 'Nov', 'Dic']

    # Ventas de huevos (diario: ultimos 30 dias)
    ventas_huevos_diarias_raw = db.session.query(
        func.date(Venta.fecha_venta).label('fecha'),
        func.sum(DetalleVenta.cantidad_huevos).label('huevos'),
        func.sum(DetalleVenta.cantidad_paneles).label('paneles'),
        func.sum(DetalleVenta.subtotal).label('total')
    ).join(DetalleVenta).filter(
        Venta.estado != 'cancelada',
        Venta.fecha_venta >= hace_30_dias
    ).group_by(
        func.date(Venta.fecha_venta)
    ).order_by(
        func.date(Venta.fecha_venta).desc()
    ).all()

    ventas_huevos_diarias = [
        {
            'fecha': item.fecha,
            'huevos': int(item.huevos or 0),
            'paneles': int(item.paneles or 0),
            'total': float(item.total or 0),
        } for item in ventas_huevos_diarias_raw
    ]

    # Ventas de huevos (mensual: anio actual)
    ventas_huevos_mensuales_raw = db.session.query(
        extract('month', Venta.fecha_venta).label('mes'),
        func.sum(DetalleVenta.cantidad_huevos).label('huevos'),
        func.sum(DetalleVenta.cantidad_paneles).label('paneles'),
        func.sum(DetalleVenta.subtotal).label('total')
    ).join(DetalleVenta).filter(
        extract('year', Venta.fecha_venta) == año_actual,
        Venta.estado != 'cancelada'
    ).group_by(
        extract('month', Venta.fecha_venta)
    ).order_by(
        extract('month', Venta.fecha_venta)
    ).all()

    ventas_huevos_mensuales = [
        {
            'mes': meses_nombres[int(item.mes) - 1],
            'huevos': int(item.huevos or 0),
            'paneles': int(item.paneles or 0),
            'total': float(item.total or 0),
        } for item in ventas_huevos_mensuales_raw
    ]

    # Produccion de huevos (diario: ultimos 30 dias)
    produccion_diaria_raw = db.session.query(
        func.date(LoteRecoleccion.fecha_recoleccion).label('fecha'),
        func.count(Huevo.id).label('huevos')
    ).join(Huevo).filter(
        LoteRecoleccion.fecha_recoleccion >= hace_30_dias,
        Huevo.roto == False
    ).group_by(
        func.date(LoteRecoleccion.fecha_recoleccion)
    ).order_by(
        func.date(LoteRecoleccion.fecha_recoleccion).desc()
    ).all()

    produccion_diaria = [
        {
            'fecha': item.fecha,
            'huevos': int(item.huevos or 0),
            'paneles': int((item.huevos or 0) // 30),
            'sobrantes': int((item.huevos or 0) % 30),
        } for item in produccion_diaria_raw
    ]

    # Produccion de huevos (mensual: anio actual)
    produccion_mensual_raw = db.session.query(
        extract('month', LoteRecoleccion.fecha_recoleccion).label('mes'),
        func.count(Huevo.id).label('huevos')
    ).join(Huevo).filter(
        extract('year', LoteRecoleccion.fecha_recoleccion) == año_actual,
        Huevo.roto == False
    ).group_by(
        extract('month', LoteRecoleccion.fecha_recoleccion)
    ).order_by(
        extract('month', LoteRecoleccion.fecha_recoleccion)
    ).all()

    produccion_mensual = [
        {
            'mes': meses_nombres[int(item.mes) - 1],
            'huevos': int(item.huevos or 0),
            'paneles': int((item.huevos or 0) // 30),
            'sobrantes': int((item.huevos or 0) % 30),
        } for item in produccion_mensual_raw
    ]

    # Tabla de disponibilidad de huevos por categoria
    disponibilidad_huevos = db.session.query(
        CategoriaHuevo.nombre,
        CategoriaHuevo.precio_venta,
        func.count(Huevo.id).label('cantidad')
    ).join(
        Huevo, CategoriaHuevo.id == Huevo.categoria_id
    ).filter(
        Huevo.roto == False,
        Huevo.vendido == False
    ).group_by(
        CategoriaHuevo.id, CategoriaHuevo.nombre, CategoriaHuevo.precio_venta
    ).order_by(
        CategoriaHuevo.peso_min
    ).all()

    # Mortalidad de gallinas por mes (anio actual)
    mortalidad_mensual_raw = db.session.query(
        extract('month', RegistroMortalidad.fecha_registro).label('mes'),
        func.sum(RegistroMortalidad.cantidad).label('cantidad')
    ).filter(
        extract('year', RegistroMortalidad.fecha_registro) == año_actual
    ).group_by(
        extract('month', RegistroMortalidad.fecha_registro)
    ).order_by(
        extract('month', RegistroMortalidad.fecha_registro)
    ).all()

    mortalidad_mensual = [0] * 12
    for item in mortalidad_mensual_raw:
        if item.mes:
            mortalidad_mensual[int(item.mes) - 1] = int(item.cantidad or 0)

    # Ventas de gallinas (ultimas 20)
    ventas_gallinas = VentaGallinas.query.order_by(VentaGallinas.fecha_venta.desc()).limit(20).all()
    
    return render_template('dashboard.html', 
                         title='Panel de Control', 
                         user=current_user,
                         datos_diarios=datos_diarios,
                         datos_mensuales=datos_mensuales,
                         distribucion_categorias=distribucion_categorias,
                         stats_generales=stats_generales,
                         ventas_huevos_diarias=ventas_huevos_diarias,
                         ventas_huevos_mensuales=ventas_huevos_mensuales,
                         produccion_diaria=produccion_diaria,
                         produccion_mensual=produccion_mensual,
                         disponibilidad_huevos=disponibilidad_huevos,
                         mortalidad_labels=meses_nombres,
                         mortalidad_mensual=mortalidad_mensual,
                         ventas_gallinas=ventas_gallinas)


@bp.route('/api/dashboard/datos/<string:periodo>')
@bp.route('/api/dashboard/datos/<string:periodo>/<int:anio>')
@bp.route('/api/dashboard/datos/<string:periodo>/<int:anio>/<int:mes>')
@login_required
def api_datos_dashboard(periodo, anio=None, mes=None):
    """API para obtener datos del dashboard en formato JSON"""
    hoy = datetime.now().date()
    
    # Si no se especifica año, usar el actual
    if anio is None:
        anio = hoy.year
    
    if periodo == 'diario':
        # Si se especifica mes, mostrar ese mes, sino los últimos 30 días
        if mes is not None:
            fecha_inicio = date(anio, mes, 1)
            if mes == 12:
                fecha_fin = date(anio + 1, 1, 1) - timedelta(days=1)
            else:
                fecha_fin = date(anio, mes + 1, 1) - timedelta(days=1)
        else:
            fecha_fin = date(anio, 12, 31) if anio != hoy.year else hoy
            fecha_inicio = fecha_fin - timedelta(days=30)
        
        # Obtener fechas únicas primero
        fechas_raw = db.session.query(
            func.date(LoteRecoleccion.fecha_recoleccion).label('fecha')
        ).join(
            Huevo, LoteRecoleccion.id == Huevo.lote_id
        ).filter(
            func.date(LoteRecoleccion.fecha_recoleccion) >= fecha_inicio,
            func.date(LoteRecoleccion.fecha_recoleccion) <= fecha_fin
        ).group_by(
            func.date(LoteRecoleccion.fecha_recoleccion)
        ).order_by(
            func.date(LoteRecoleccion.fecha_recoleccion)
        ).all()
        
        # Crear datos completos
        datos = []
        for fecha_obj in fechas_raw:
            fecha = fecha_obj.fecha
            
            total_huevos = db.session.query(func.count(Huevo.id)).join(LoteRecoleccion).filter(
                func.date(LoteRecoleccion.fecha_recoleccion) == fecha
            ).scalar() or 0
            
            huevos_buenos = db.session.query(func.count(Huevo.id)).join(LoteRecoleccion).filter(
                func.date(LoteRecoleccion.fecha_recoleccion) == fecha,
                Huevo.roto == False
            ).scalar() or 0
            
            huevos_rotos = db.session.query(func.count(Huevo.id)).join(LoteRecoleccion).filter(
                func.date(LoteRecoleccion.fecha_recoleccion) == fecha,
                Huevo.roto == True
            ).scalar() or 0
            
            peso_total = db.session.query(func.sum(Huevo.peso)).join(LoteRecoleccion).filter(
                func.date(LoteRecoleccion.fecha_recoleccion) == fecha,
                Huevo.roto == False
            ).scalar() or 0
            
            class DatoDiario:
                def __init__(self, fecha, total, buenos, rotos, peso):
                    self.fecha = fecha
                    self.total_huevos = total
                    self.huevos_buenos = buenos
                    self.huevos_rotos = rotos
                    self.peso_total = peso
                    
            datos.append(DatoDiario(fecha, total_huevos, huevos_buenos, huevos_rotos, peso_total))
        
        return jsonify({
            'labels': [dato.fecha.strftime('%d/%m') for dato in datos],
            'datasets': [
                {
                    'label': 'Huevos Buenos',
                    'data': [int(dato.huevos_buenos or 0) for dato in datos],
                    'backgroundColor': 'rgba(40, 167, 69, 0.8)',
                    'borderColor': 'rgba(40, 167, 69, 1)'
                },
                {
                    'label': 'Huevos Rotos',
                    'data': [int(dato.huevos_rotos or 0) for dato in datos],
                    'backgroundColor': 'rgba(220, 53, 69, 0.8)',
                    'borderColor': 'rgba(220, 53, 69, 1)'
                },
                {
                    'label': 'Peso Total (g)',
                    'data': [round(float(dato.peso_total or 0), 1) for dato in datos],
                    'backgroundColor': 'rgba(0, 123, 255, 0.8)',
                    'borderColor': 'rgba(0, 123, 255, 1)',
                    'yAxisID': 'y1'
                }
            ],
            'tabla': [
                {
                    'fecha': dato.fecha.strftime('%d/%m/%Y'),
                    'total_huevos': int(dato.total_huevos or 0),
                    'huevos_buenos': int(dato.huevos_buenos or 0),
                    'huevos_rotos': int(dato.huevos_rotos or 0),
                    'peso_total': round(float(dato.peso_total or 0), 1)
                } for dato in datos
            ]
        })
        
    elif periodo == 'mensual':
        meses = ['Ene', 'Feb', 'Mar', 'Abr', 'May', 'Jun', 
                'Jul', 'Ago', 'Sep', 'Oct', 'Nov', 'Dic']
        
        # Obtener meses únicos primero
        meses_raw = db.session.query(
            extract('month', LoteRecoleccion.fecha_recoleccion).label('mes')
        ).join(
            Huevo, LoteRecoleccion.id == Huevo.lote_id
        ).filter(
            extract('year', LoteRecoleccion.fecha_recoleccion) == anio
        ).group_by(
            extract('month', LoteRecoleccion.fecha_recoleccion)
        ).order_by(
            extract('month', LoteRecoleccion.fecha_recoleccion)
        ).all()
        
        # Crear datos completos
        datos = []
        for mes_obj in meses_raw:
            mes = mes_obj.mes
            
            total_huevos = db.session.query(func.count(Huevo.id)).join(LoteRecoleccion).filter(
                extract('year', LoteRecoleccion.fecha_recoleccion) == anio,
                extract('month', LoteRecoleccion.fecha_recoleccion) == mes
            ).scalar() or 0
            
            huevos_buenos = db.session.query(func.count(Huevo.id)).join(LoteRecoleccion).filter(
                extract('year', LoteRecoleccion.fecha_recoleccion) == anio,
                extract('month', LoteRecoleccion.fecha_recoleccion) == mes,
                Huevo.roto == False
            ).scalar() or 0
            
            huevos_rotos = db.session.query(func.count(Huevo.id)).join(LoteRecoleccion).filter(
                extract('year', LoteRecoleccion.fecha_recoleccion) == anio,
                extract('month', LoteRecoleccion.fecha_recoleccion) == mes,
                Huevo.roto == True
            ).scalar() or 0
            
            peso_total = db.session.query(func.sum(Huevo.peso)).join(LoteRecoleccion).filter(
                extract('year', LoteRecoleccion.fecha_recoleccion) == anio,
                extract('month', LoteRecoleccion.fecha_recoleccion) == mes,
                Huevo.roto == False
            ).scalar() or 0
            
            class DatoMensual:
                def __init__(self, mes, total, buenos, rotos, peso):
                    self.mes = mes
                    self.total_huevos = total
                    self.huevos_buenos = buenos
                    self.huevos_rotos = rotos
                    self.peso_total = peso
                    
            datos.append(DatoMensual(mes, total_huevos, huevos_buenos, huevos_rotos, peso_total))
        
        # Crear array con todos los meses (incluyendo los que no tienen datos)
        datos_completos = []
        for i in range(1, 13):
            dato_mes = next((d for d in datos if d.mes == i), None)
            if dato_mes:
                datos_completos.append(dato_mes)
            else:
                # Mes sin datos
                class DatoVacio:
                    def __init__(self, mes):
                        self.mes = mes
                        self.total_huevos = 0
                        self.huevos_buenos = 0
                        self.huevos_rotos = 0
                        self.peso_total = 0
                datos_completos.append(DatoVacio(i))
        
        return jsonify({
            'labels': [meses[dato.mes - 1] for dato in datos_completos],
            'datasets': [
                {
                    'label': 'Huevos Buenos',
                    'data': [int(dato.huevos_buenos or 0) for dato in datos_completos],
                    'backgroundColor': 'rgba(40, 167, 69, 0.8)',
                    'borderColor': 'rgba(40, 167, 69, 1)'
                },
                {
                    'label': 'Huevos Rotos',
                    'data': [int(dato.huevos_rotos or 0) for dato in datos_completos],
                    'backgroundColor': 'rgba(220, 53, 69, 0.8)',
                    'borderColor': 'rgba(220, 53, 69, 1)'
                },
                {
                    'label': 'Peso Total (g)',
                    'data': [round(float(dato.peso_total or 0), 1) for dato in datos_completos],
                    'backgroundColor': 'rgba(0, 123, 255, 0.8)',
                    'borderColor': 'rgba(0, 123, 255, 1)',
                    'yAxisID': 'y1'
                }
            ],
            'tabla': [
                {
                    'mes': meses[dato.mes-1],
                    'total_huevos': int(dato.total_huevos or 0),
                    'huevos_buenos': int(dato.huevos_buenos or 0),
                    'huevos_rotos': int(dato.huevos_rotos or 0),
                    'peso_total': round(float(dato.peso_total or 0), 1)
                } for dato in datos_completos
            ]
        })
    
    return jsonify({'error': 'Periodo no válido'}), 400


@bp.route('/api/dashboard/datos-categorias/<string:periodo>')
@bp.route('/api/dashboard/datos-categorias/<string:periodo>/<int:anio>')
@bp.route('/api/dashboard/datos-categorias/<string:periodo>/<int:anio>/<int:mes>')
@login_required
def api_datos_categorias(periodo, anio=None, mes=None):
    """API para obtener datos del dashboard con detalle por categorías"""
    hoy = datetime.now().date()
    
    # Si no se especifica año, usar el actual
    if anio is None:
        anio = hoy.year
    
    if periodo == 'diario':
        # Si se especifica mes, mostrar ese mes, sino los últimos 30 días
        if mes is not None:
            fecha_inicio = date(anio, mes, 1)
            if mes == 12:
                fecha_fin = date(anio + 1, 1, 1) - timedelta(days=1)
            else:
                fecha_fin = date(anio, mes + 1, 1) - timedelta(days=1)
        else:
            fecha_fin = date(anio, 12, 31) if anio != hoy.year else hoy
            fecha_inicio = fecha_fin - timedelta(days=30)
        
        # Obtener fechas únicas
        fechas = db.session.query(
            func.date(LoteRecoleccion.fecha_recoleccion).label('fecha')
        ).join(Huevo).filter(
            func.date(LoteRecoleccion.fecha_recoleccion) >= fecha_inicio,
            func.date(LoteRecoleccion.fecha_recoleccion) <= fecha_fin
        ).group_by(func.date(LoteRecoleccion.fecha_recoleccion)).order_by(
            func.date(LoteRecoleccion.fecha_recoleccion)
        ).all()
        
        # Obtener categorías
        categorias = CategoriaHuevo.query.filter(CategoriaHuevo.activo == True).all()
        
        datos_tabla = []
        for fecha_obj in fechas:
            fecha = fecha_obj.fecha
            fila = {
                'fecha': fecha.strftime('%d/%m/%Y'),
                'total_huevos': 0,
                'huevos_buenos': 0,
                'huevos_rotos': 0,
                'peso_total': 0,
                'categorias': {}
            }
            
            # Datos generales
            total = db.session.query(func.count(Huevo.id)).join(LoteRecoleccion).filter(
                func.date(LoteRecoleccion.fecha_recoleccion) == fecha
            ).scalar() or 0
            
            buenos = db.session.query(func.count(Huevo.id)).join(LoteRecoleccion).filter(
                func.date(LoteRecoleccion.fecha_recoleccion) == fecha,
                Huevo.roto == False
            ).scalar() or 0
            
            rotos = db.session.query(func.count(Huevo.id)).join(LoteRecoleccion).filter(
                func.date(LoteRecoleccion.fecha_recoleccion) == fecha,
                Huevo.roto == True
            ).scalar() or 0
            
            peso = db.session.query(func.sum(Huevo.peso)).join(LoteRecoleccion).filter(
                func.date(LoteRecoleccion.fecha_recoleccion) == fecha,
                Huevo.roto == False
            ).scalar() or 0
            
            fila.update({
                'total_huevos': total,
                'huevos_buenos': buenos,
                'huevos_rotos': rotos,
                'peso_total': round(float(peso), 1)
            })
            
            # Datos por categoría
            for categoria in categorias:
                cantidad = db.session.query(func.count(Huevo.id)).join(LoteRecoleccion).filter(
                    func.date(LoteRecoleccion.fecha_recoleccion) == fecha,
                    Huevo.categoria_id == categoria.id,
                    Huevo.roto == False
                ).scalar() or 0
                
                peso_cat = db.session.query(func.sum(Huevo.peso)).join(LoteRecoleccion).filter(
                    func.date(LoteRecoleccion.fecha_recoleccion) == fecha,
                    Huevo.categoria_id == categoria.id,
                    Huevo.roto == False
                ).scalar() or 0
                
                fila['categorias'][categoria.nombre] = {
                    'cantidad': cantidad,
                    'peso': round(float(peso_cat), 1)
                }
            
            datos_tabla.append(fila)
        
        return jsonify({
            'tabla': datos_tabla,
            'categorias': [cat.nombre for cat in categorias]
        })
    
    elif periodo == 'mensual':
        meses = ['Ene', 'Feb', 'Mar', 'Abr', 'May', 'Jun', 
                'Jul', 'Ago', 'Sep', 'Oct', 'Nov', 'Dic']
        
        # Obtener meses con datos
        meses_con_datos = db.session.query(
            extract('month', LoteRecoleccion.fecha_recoleccion).label('mes')
        ).join(Huevo).filter(
            extract('year', LoteRecoleccion.fecha_recoleccion) == anio
        ).group_by(extract('month', LoteRecoleccion.fecha_recoleccion)).all()
        
        # Obtener categorías
        categorias = CategoriaHuevo.query.filter(CategoriaHuevo.activo == True).all()
        
        datos_tabla = []
        for i in range(1, 13):  # Todos los meses del año
            mes = i
            fila = {
                'mes': meses[mes - 1],
                'total_huevos': 0,
                'huevos_buenos': 0,
                'huevos_rotos': 0,
                'peso_total': 0,
                'categorias': {}
            }
            
            # Solo procesar si hay datos para este mes
            if any(m.mes == mes for m in meses_con_datos):
                # Datos generales
                total = db.session.query(func.count(Huevo.id)).join(LoteRecoleccion).filter(
                    extract('year', LoteRecoleccion.fecha_recoleccion) == anio,
                    extract('month', LoteRecoleccion.fecha_recoleccion) == mes
                ).scalar() or 0
                
                buenos = db.session.query(func.count(Huevo.id)).join(LoteRecoleccion).filter(
                    extract('year', LoteRecoleccion.fecha_recoleccion) == anio,
                    extract('month', LoteRecoleccion.fecha_recoleccion) == mes,
                    Huevo.roto == False
                ).scalar() or 0
                
                rotos = db.session.query(func.count(Huevo.id)).join(LoteRecoleccion).filter(
                    extract('year', LoteRecoleccion.fecha_recoleccion) == anio,
                    extract('month', LoteRecoleccion.fecha_recoleccion) == mes,
                    Huevo.roto == True
                ).scalar() or 0
                
                peso = db.session.query(func.sum(Huevo.peso)).join(LoteRecoleccion).filter(
                    extract('year', LoteRecoleccion.fecha_recoleccion) == anio,
                    extract('month', LoteRecoleccion.fecha_recoleccion) == mes,
                    Huevo.roto == False
                ).scalar() or 0
                
                fila.update({
                    'total_huevos': total,
                    'huevos_buenos': buenos,
                    'huevos_rotos': rotos,
                    'peso_total': round(float(peso), 1)
                })
                
                # Datos por categoría
                for categoria in categorias:
                    cantidad = db.session.query(func.count(Huevo.id)).join(LoteRecoleccion).filter(
                        extract('year', LoteRecoleccion.fecha_recoleccion) == anio,
                        extract('month', LoteRecoleccion.fecha_recoleccion) == mes,
                        Huevo.categoria_id == categoria.id,
                        Huevo.roto == False
                    ).scalar() or 0
                    
                    peso_cat = db.session.query(func.sum(Huevo.peso)).join(LoteRecoleccion).filter(
                        extract('year', LoteRecoleccion.fecha_recoleccion) == anio,
                        extract('month', LoteRecoleccion.fecha_recoleccion) == mes,
                        Huevo.categoria_id == categoria.id,
                        Huevo.roto == False
                    ).scalar() or 0
                    
                    fila['categorias'][categoria.nombre] = {
                        'cantidad': cantidad,
                        'peso': round(float(peso_cat), 1)
                    }
            else:
                # Mes sin datos - inicializar categorías vacías
                for categoria in categorias:
                    fila['categorias'][categoria.nombre] = {
                        'cantidad': 0,
                        'peso': 0
                    }
            
            datos_tabla.append(fila)
        
        return jsonify({
            'tabla': datos_tabla,
            'categorias': [cat.nombre for cat in categorias]
        })
    
    elif periodo == 'anual':
        # Vista diaria para todo el año actual
        inicio_año = datetime(hoy.year, 1, 1).date()
        
        # Obtener fechas con datos
        fechas = db.session.query(
            func.date(LoteRecoleccion.fecha_recoleccion).label('fecha')
        ).join(Huevo).filter(
            extract('year', LoteRecoleccion.fecha_recoleccion) == hoy.year
        ).group_by(func.date(LoteRecoleccion.fecha_recoleccion)).order_by(
            func.date(LoteRecoleccion.fecha_recoleccion).desc()
        ).all()
        
        # Obtener categorías
        categorias = CategoriaHuevo.query.filter(CategoriaHuevo.activo == True).all()
        
        datos_tabla = []
        for fecha_obj in fechas:
            fecha = fecha_obj.fecha
            fila = {
                'fecha': fecha.strftime('%d/%m/%Y'),
                'total_huevos': 0,
                'huevos_buenos': 0,
                'huevos_rotos': 0,
                'peso_total': 0,
                'categorias': {}
            }
            
            # Datos generales (igual que el diario)
            total = db.session.query(func.count(Huevo.id)).join(LoteRecoleccion).filter(
                func.date(LoteRecoleccion.fecha_recoleccion) == fecha
            ).scalar() or 0
            
            buenos = db.session.query(func.count(Huevo.id)).join(LoteRecoleccion).filter(
                func.date(LoteRecoleccion.fecha_recoleccion) == fecha,
                Huevo.roto == False
            ).scalar() or 0
            
            rotos = db.session.query(func.count(Huevo.id)).join(LoteRecoleccion).filter(
                func.date(LoteRecoleccion.fecha_recoleccion) == fecha,
                Huevo.roto == True
            ).scalar() or 0
            
            peso = db.session.query(func.sum(Huevo.peso)).join(LoteRecoleccion).filter(
                func.date(LoteRecoleccion.fecha_recoleccion) == fecha,
                Huevo.roto == False
            ).scalar() or 0
            
            fila.update({
                'total_huevos': total,
                'huevos_buenos': buenos,
                'huevos_rotos': rotos,
                'peso_total': round(float(peso), 1)
            })
            
            # Datos por categoría
            for categoria in categorias:
                cantidad = db.session.query(func.count(Huevo.id)).join(LoteRecoleccion).filter(
                    func.date(LoteRecoleccion.fecha_recoleccion) == fecha,
                    Huevo.categoria_id == categoria.id,
                    Huevo.roto == False
                ).scalar() or 0
                
                peso_cat = db.session.query(func.sum(Huevo.peso)).join(LoteRecoleccion).filter(
                    func.date(LoteRecoleccion.fecha_recoleccion) == fecha,
                    Huevo.categoria_id == categoria.id,
                    Huevo.roto == False
                ).scalar() or 0
                
                fila['categorias'][categoria.nombre] = {
                    'cantidad': cantidad,
                    'peso': round(float(peso_cat), 1)
                }
            
            datos_tabla.append(fila)
        
        return jsonify({
            'tabla': datos_tabla,
            'categorias': [cat.nombre for cat in categorias]
        })
    
    return jsonify({'error': 'Periodo no válido'}), 400


@bp.route('/api/dashboard/anos-disponibles')
@login_required
def api_anos_disponibles():
    """API para obtener los años que tienen datos"""
    years = db.session.query(
        extract('year', LoteRecoleccion.fecha_recoleccion).label('year')
    ).join(Huevo).group_by(
        extract('year', LoteRecoleccion.fecha_recoleccion)
    ).order_by(
        extract('year', LoteRecoleccion.fecha_recoleccion).desc()
    ).all()
    
    return jsonify({
        'anos': [int(year.year) for year in years]
    })


@bp.route('/api/dashboard/meses-disponibles/<int:anio>')
@login_required
def api_meses_disponibles(anio):
    """API para obtener los meses que tienen datos en un año específico"""
    meses = db.session.query(
        extract('month', LoteRecoleccion.fecha_recoleccion).label('mes')
    ).join(Huevo).filter(
        extract('year', LoteRecoleccion.fecha_recoleccion) == anio
    ).group_by(
        extract('month', LoteRecoleccion.fecha_recoleccion)
    ).order_by(
        extract('month', LoteRecoleccion.fecha_recoleccion)
    ).all()
    
    meses_nombres = ['Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
                    'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']
    
    return jsonify({
        'meses': [
            {'numero': int(mes.mes), 'nombre': meses_nombres[int(mes.mes) - 1]} 
            for mes in meses
        ]
    })


@bp.route('/api/dashboard/categorias-stats/<string:periodo>')
@bp.route('/api/dashboard/categorias-stats/<string:periodo>/<int:anio>')
@bp.route('/api/dashboard/categorias-stats/<string:periodo>/<int:anio>/<int:mes>')
@login_required
def api_categorias_stats(periodo, anio=None, mes=None):
    """API para obtener estadísticas de categorías para gráficos"""
    hoy = datetime.now().date()
    
    # Si no se especifica año, usar el actual
    if anio is None:
        anio = hoy.year
    
    # Obtener categorías activas
    categorias = CategoriaHuevo.query.filter(CategoriaHuevo.activo == True).order_by(CategoriaHuevo.peso_min).all()
    
    if periodo == 'diario':
        # Si se especifica mes, mostrar ese mes, sino los últimos 30 días
        if mes is not None:
            fecha_inicio = date(anio, mes, 1)
            if mes == 12:
                fecha_fin = date(anio + 1, 1, 1) - timedelta(days=1)
            else:
                fecha_fin = date(anio, mes + 1, 1) - timedelta(days=1)
        else:
            fecha_fin = date(anio, 12, 31) if anio != hoy.year else hoy
            fecha_inicio = fecha_fin - timedelta(days=30)
        
        # Obtener fechas únicas
        fechas = db.session.query(
            func.date(LoteRecoleccion.fecha_recoleccion).label('fecha')
        ).join(Huevo).filter(
            func.date(LoteRecoleccion.fecha_recoleccion) >= fecha_inicio,
            func.date(LoteRecoleccion.fecha_recoleccion) <= fecha_fin
        ).group_by(func.date(LoteRecoleccion.fecha_recoleccion)).order_by(
            func.date(LoteRecoleccion.fecha_recoleccion)
        ).all()
        
        # Datos para gráfico de evolución
        evolucion_data = {
            'labels': [fecha.fecha.strftime('%d/%m') for fecha in fechas],
            'datasets': []
        }
        
        colores = [
            'rgba(255, 99, 132, 0.8)',   # Rojo
            'rgba(54, 162, 235, 0.8)',   # Azul  
            'rgba(255, 205, 86, 0.8)',   # Amarillo
            'rgba(75, 192, 192, 0.8)',   # Verde agua
            'rgba(153, 102, 255, 0.8)',  # Púrpura
            'rgba(255, 159, 64, 0.8)',   # Naranja
        ]
        
        # Datos totales por categoría para gráfico de barras y torta
        totales_categorias = {}
        
        for i, categoria in enumerate(categorias):
            # Datos por fecha para evolución
            datos_fecha = []
            total_categoria = 0
            
            for fecha_obj in fechas:
                fecha = fecha_obj.fecha
                cantidad = db.session.query(func.count(Huevo.id)).join(LoteRecoleccion).filter(
                    func.date(LoteRecoleccion.fecha_recoleccion) == fecha,
                    Huevo.categoria_id == categoria.id,
                    Huevo.roto == False
                ).scalar() or 0
                
                datos_fecha.append(cantidad)
                total_categoria += cantidad
            
            # Dataset para evolución
            evolucion_data['datasets'].append({
                'label': categoria.nombre,
                'data': datos_fecha,
                'borderColor': colores[i % len(colores)],
                'backgroundColor': colores[i % len(colores)].replace('0.8', '0.2'),
                'fill': False,
                'tension': 0.3
            })
            
            totales_categorias[categoria.nombre] = total_categoria
        
    elif periodo == 'mensual':
        # Obtener meses con datos
        meses_raw = db.session.query(
            extract('month', LoteRecoleccion.fecha_recoleccion).label('mes')
        ).join(Huevo).filter(
            extract('year', LoteRecoleccion.fecha_recoleccion) == anio
        ).group_by(
            extract('month', LoteRecoleccion.fecha_recoleccion)
        ).order_by(
            extract('month', LoteRecoleccion.fecha_recoleccion)
        ).all()
        
        meses_nombres = ['Ene', 'Feb', 'Mar', 'Abr', 'May', 'Jun', 
                        'Jul', 'Ago', 'Sep', 'Oct', 'Nov', 'Dic']
        
        # Crear lista completa de meses (1-12)
        todos_meses = list(range(1, 13))
        meses_con_datos = [m.mes for m in meses_raw]
        
        # Datos para gráfico de evolución
        evolucion_data = {
            'labels': [meses_nombres[mes - 1] for mes in todos_meses],
            'datasets': []
        }
        
        colores = [
            'rgba(255, 99, 132, 0.8)',   # Rojo
            'rgba(54, 162, 235, 0.8)',   # Azul  
            'rgba(255, 205, 86, 0.8)',   # Amarillo
            'rgba(75, 192, 192, 0.8)',   # Verde agua
            'rgba(153, 102, 255, 0.8)',  # Púrpura
            'rgba(255, 159, 64, 0.8)',   # Naranja
        ]
        
        # Datos totales por categoría
        totales_categorias = {}
        
        for i, categoria in enumerate(categorias):
            # Datos por mes para evolución
            datos_mes = []
            total_categoria = 0
            
            for mes in todos_meses:
                if mes in meses_con_datos:
                    cantidad = db.session.query(func.count(Huevo.id)).join(LoteRecoleccion).filter(
                        extract('year', LoteRecoleccion.fecha_recoleccion) == anio,
                        extract('month', LoteRecoleccion.fecha_recoleccion) == mes,
                        Huevo.categoria_id == categoria.id,
                        Huevo.roto == False
                    ).scalar() or 0
                else:
                    cantidad = 0
                
                datos_mes.append(cantidad)
                total_categoria += cantidad
            
            # Dataset para evolución
            evolucion_data['datasets'].append({
                'label': categoria.nombre,
                'data': datos_mes,
                'borderColor': colores[i % len(colores)],
                'backgroundColor': colores[i % len(colores)].replace('0.8', '0.2'),
                'fill': False,
                'tension': 0.3
            })
            
            totales_categorias[categoria.nombre] = total_categoria
    
    # Preparar datos para gráfico de barras
    barras_data = {
        'labels': list(totales_categorias.keys()),
        'datasets': [{
            'label': 'Huevos por Categoría',
            'data': list(totales_categorias.values()),
            'backgroundColor': [
                'rgba(255, 99, 132, 0.8)',
                'rgba(54, 162, 235, 0.8)',
                'rgba(255, 205, 86, 0.8)',
                'rgba(75, 192, 192, 0.8)',
                'rgba(153, 102, 255, 0.8)',
                'rgba(255, 159, 64, 0.8)',
            ],
            'borderColor': [
                'rgba(255, 99, 132, 1)',
                'rgba(54, 162, 235, 1)',
                'rgba(255, 205, 86, 1)',
                'rgba(75, 192, 192, 1)',
                'rgba(153, 102, 255, 1)',
                'rgba(255, 159, 64, 1)',
            ],
            'borderWidth': 1
        }]
    }
    
    # Preparar datos para gráfico de torta
    torta_data = {
        'labels': list(totales_categorias.keys()),
        'datasets': [{
            'data': list(totales_categorias.values()),
            'backgroundColor': [
                'rgba(255, 99, 132, 0.8)',
                'rgba(54, 162, 235, 0.8)',
                'rgba(255, 205, 86, 0.8)',
                'rgba(75, 192, 192, 0.8)',
                'rgba(153, 102, 255, 0.8)',
                'rgba(255, 159, 64, 0.8)',
            ],
            'borderColor': [
                'rgba(255, 99, 132, 1)',
                'rgba(54, 162, 235, 1)',
                'rgba(255, 205, 86, 1)',
                'rgba(75, 192, 192, 1)',
                'rgba(153, 102, 255, 1)',
                'rgba(255, 159, 64, 1)',
            ],
            'borderWidth': 2
        }]
    }
    
    return jsonify({
        'barras': barras_data,
        'torta': torta_data,
        'evolucion': evolucion_data
    })


@bp.route('/api/dashboard/export/<string:reporte>')
@login_required
def export_dashboard(reporte):
    """Exportar reportes del dashboard en CSV"""
    periodo = request.args.get('periodo', 'diario')
    anio = request.args.get('anio', type=int)
    mes = request.args.get('mes', type=int)
    hoy = datetime.now().date()
    if anio is None:
        anio = hoy.year
    meses_nombres = ['Ene', 'Feb', 'Mar', 'Abr', 'May', 'Jun',
                     'Jul', 'Ago', 'Sep', 'Oct', 'Nov', 'Dic']

    output = io.StringIO()
    writer = csv.writer(output)

    if reporte == 'inventario-resumen':
        # Reutiliza la misma logica del dashboard para resumen por fecha/mes
        if periodo == 'diario':
            if mes is not None:
                fecha_inicio = date(anio, mes, 1)
                fecha_fin = date(anio + 1, 1, 1) - timedelta(days=1) if mes == 12 else date(anio, mes + 1, 1) - timedelta(days=1)
            else:
                fecha_inicio = date(anio, 1, 1)
                fecha_fin = hoy if anio == hoy.year else date(anio, 12, 31)
            fechas = db.session.query(func.date(LoteRecoleccion.fecha_recoleccion).label('fecha')).join(
                Huevo, LoteRecoleccion.id == Huevo.lote_id
            ).filter(
                func.date(LoteRecoleccion.fecha_recoleccion) >= fecha_inicio,
                func.date(LoteRecoleccion.fecha_recoleccion) <= fecha_fin
            ).group_by(func.date(LoteRecoleccion.fecha_recoleccion)).order_by(func.date(LoteRecoleccion.fecha_recoleccion)).all()
            writer.writerow(['Fecha', 'Total Huevos', 'Buenos', 'Rotos', 'Peso Total (g)'])
            for fecha_obj in fechas:
                fecha = fecha_obj.fecha
                total = db.session.query(func.count(Huevo.id)).join(LoteRecoleccion).filter(
                    func.date(LoteRecoleccion.fecha_recoleccion) == fecha
                ).scalar() or 0
                buenos = db.session.query(func.count(Huevo.id)).join(LoteRecoleccion).filter(
                    func.date(LoteRecoleccion.fecha_recoleccion) == fecha,
                    Huevo.roto == False
                ).scalar() or 0
                rotos = db.session.query(func.count(Huevo.id)).join(LoteRecoleccion).filter(
                    func.date(LoteRecoleccion.fecha_recoleccion) == fecha,
                    Huevo.roto == True
                ).scalar() or 0
                peso = db.session.query(func.sum(Huevo.peso)).join(LoteRecoleccion).filter(
                    func.date(LoteRecoleccion.fecha_recoleccion) == fecha,
                    Huevo.roto == False
                ).scalar() or 0
                writer.writerow([fecha.strftime('%d/%m/%Y'), int(total), int(buenos), int(rotos), round(float(peso or 0), 1)])
        else:
            writer.writerow(['Mes', 'Total Huevos', 'Buenos', 'Rotos', 'Peso Total (g)'])
            for i in range(1, 13):
                total = db.session.query(func.count(Huevo.id)).join(LoteRecoleccion).filter(
                    extract('year', LoteRecoleccion.fecha_recoleccion) == anio,
                    extract('month', LoteRecoleccion.fecha_recoleccion) == i
                ).scalar() or 0
                buenos = db.session.query(func.count(Huevo.id)).join(LoteRecoleccion).filter(
                    extract('year', LoteRecoleccion.fecha_recoleccion) == anio,
                    extract('month', LoteRecoleccion.fecha_recoleccion) == i,
                    Huevo.roto == False
                ).scalar() or 0
                rotos = db.session.query(func.count(Huevo.id)).join(LoteRecoleccion).filter(
                    extract('year', LoteRecoleccion.fecha_recoleccion) == anio,
                    extract('month', LoteRecoleccion.fecha_recoleccion) == i,
                    Huevo.roto == True
                ).scalar() or 0
                peso = db.session.query(func.sum(Huevo.peso)).join(LoteRecoleccion).filter(
                    extract('year', LoteRecoleccion.fecha_recoleccion) == anio,
                    extract('month', LoteRecoleccion.fecha_recoleccion) == i,
                    Huevo.roto == False
                ).scalar() or 0
                writer.writerow([meses_nombres[i - 1], int(total), int(buenos), int(rotos), round(float(peso or 0), 1)])

    elif reporte == 'ventas-huevos':
        if periodo == 'diario':
            if mes is not None:
                fecha_inicio = date(anio, mes, 1)
                fecha_fin = date(anio + 1, 1, 1) - timedelta(days=1) if mes == 12 else date(anio, mes + 1, 1) - timedelta(days=1)
            else:
                fecha_inicio = date(anio, 1, 1)
                fecha_fin = hoy if anio == hoy.year else date(anio, 12, 31)
            rows = db.session.query(
                func.date(Venta.fecha_venta).label('fecha'),
                func.sum(DetalleVenta.cantidad_huevos).label('huevos'),
                func.sum(DetalleVenta.cantidad_paneles).label('paneles'),
                func.sum(DetalleVenta.subtotal).label('total')
            ).join(DetalleVenta).filter(
                Venta.estado != 'cancelada',
                Venta.fecha_venta >= fecha_inicio,
                Venta.fecha_venta <= fecha_fin
            ).group_by(func.date(Venta.fecha_venta)).order_by(func.date(Venta.fecha_venta)).all()
            writer.writerow(['Fecha', 'Huevos', 'Paneles', 'Total'])
            for row in rows:
                writer.writerow([row.fecha.strftime('%d/%m/%Y'), int(row.huevos or 0), int(row.paneles or 0), round(float(row.total or 0), 2)])
        else:
            rows = db.session.query(
                extract('month', Venta.fecha_venta).label('mes'),
                func.sum(DetalleVenta.cantidad_huevos).label('huevos'),
                func.sum(DetalleVenta.cantidad_paneles).label('paneles'),
                func.sum(DetalleVenta.subtotal).label('total')
            ).join(DetalleVenta).filter(
                extract('year', Venta.fecha_venta) == anio,
                Venta.estado != 'cancelada'
            ).group_by(extract('month', Venta.fecha_venta)).order_by(extract('month', Venta.fecha_venta)).all()
            writer.writerow(['Mes', 'Huevos', 'Paneles', 'Total'])
            for row in rows:
                writer.writerow([meses_nombres[int(row.mes) - 1], int(row.huevos or 0), int(row.paneles or 0), round(float(row.total or 0), 2)])

    elif reporte == 'produccion-huevos':
        if periodo == 'diario':
            if mes is not None:
                fecha_inicio = date(anio, mes, 1)
                fecha_fin = date(anio + 1, 1, 1) - timedelta(days=1) if mes == 12 else date(anio, mes + 1, 1) - timedelta(days=1)
            else:
                fecha_inicio = date(anio, 1, 1)
                fecha_fin = hoy if anio == hoy.year else date(anio, 12, 31)
            rows = db.session.query(
                func.date(LoteRecoleccion.fecha_recoleccion).label('fecha'),
                func.count(Huevo.id).label('huevos')
            ).join(Huevo).filter(
                LoteRecoleccion.fecha_recoleccion >= fecha_inicio,
                LoteRecoleccion.fecha_recoleccion <= fecha_fin,
                Huevo.roto == False
            ).group_by(func.date(LoteRecoleccion.fecha_recoleccion)).order_by(func.date(LoteRecoleccion.fecha_recoleccion)).all()
            writer.writerow(['Fecha', 'Huevos', 'Paneles', 'Sobrantes'])
            for row in rows:
                huevos = int(row.huevos or 0)
                writer.writerow([row.fecha.strftime('%d/%m/%Y'), huevos, huevos // 30, huevos % 30])
        else:
            rows = db.session.query(
                extract('month', LoteRecoleccion.fecha_recoleccion).label('mes'),
                func.count(Huevo.id).label('huevos')
            ).join(Huevo).filter(
                extract('year', LoteRecoleccion.fecha_recoleccion) == anio,
                Huevo.roto == False
            ).group_by(extract('month', LoteRecoleccion.fecha_recoleccion)).order_by(extract('month', LoteRecoleccion.fecha_recoleccion)).all()
            writer.writerow(['Mes', 'Huevos', 'Paneles', 'Sobrantes'])
            for row in rows:
                huevos = int(row.huevos or 0)
                writer.writerow([meses_nombres[int(row.mes) - 1], huevos, huevos // 30, huevos % 30])

    elif reporte == 'inventario-disponible':
        rows = db.session.query(
            CategoriaHuevo.nombre,
            CategoriaHuevo.precio_venta,
            func.count(Huevo.id).label('cantidad')
        ).join(Huevo, CategoriaHuevo.id == Huevo.categoria_id).filter(
            Huevo.roto == False,
            Huevo.vendido == False
        ).group_by(CategoriaHuevo.id, CategoriaHuevo.nombre, CategoriaHuevo.precio_venta).order_by(CategoriaHuevo.peso_min).all()
        writer.writerow(['Categoria', 'Huevos', 'Paneles', 'Sobrantes', 'Precio'])
        for row in rows:
            huevos = int(row.cantidad or 0)
            writer.writerow([row.nombre, huevos, huevos // 30, huevos % 30, round(float(row.precio_venta or 0), 2)])

    elif reporte == 'mortalidad-gallinas':
        writer.writerow(['Mes', 'Cantidad'])
        for i in range(1, 13):
            cantidad = db.session.query(func.sum(RegistroMortalidad.cantidad)).filter(
                extract('year', RegistroMortalidad.fecha_registro) == anio,
                extract('month', RegistroMortalidad.fecha_registro) == i
            ).scalar() or 0
            writer.writerow([meses_nombres[i - 1], int(cantidad)])

    elif reporte == 'ventas-gallinas':
        rows = VentaGallinas.query.order_by(VentaGallinas.fecha_venta.desc()).all()
        writer.writerow(['Fecha', 'Cantidad', 'Precio Unitario', 'Precio Total', 'Comprador', 'Observaciones'])
        for row in rows:
            writer.writerow([
                row.fecha_venta.strftime('%d/%m/%Y'),
                int(row.cantidad or 0),
                round(float(row.precio_unitario or 0), 2),
                round(float(row.precio_total or 0), 2),
                row.comprador or '',
                row.observaciones or ''
            ])
    else:
        return jsonify({'error': 'Reporte no valido'}), 400

    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'text/csv; charset=utf-8'
    response.headers['Content-Disposition'] = f'attachment; filename={reporte}_{periodo}_{anio}.csv'
    return response


@bp.route('/api/dashboard/reportes')
@login_required
def api_reportes_dashboard():
    """API para reportes adicionales del dashboard con filtros"""
    periodo = request.args.get('periodo', 'diario')
    anio = request.args.get('anio', type=int)
    mes = request.args.get('mes', type=int)
    return jsonify(_build_reportes_data(periodo, anio, mes))

def _build_reportes_data(periodo, anio, mes):
    hoy = datetime.now().date()
    if anio is None:
        anio = hoy.year

    meses_nombres = ['Ene', 'Feb', 'Mar', 'Abr', 'May', 'Jun',
                     'Jul', 'Ago', 'Sep', 'Oct', 'Nov', 'Dic']

    def rango_diario():
        if mes:
            fecha_inicio = date(anio, mes, 1)
            if mes == 12:
                fecha_fin = date(anio + 1, 1, 1) - timedelta(days=1)
            else:
                fecha_fin = date(anio, mes + 1, 1) - timedelta(days=1)
        else:
            # Si hay anio seleccionado, usar todo el anio (hasta hoy si es el actual)
            fecha_inicio = date(anio, 1, 1)
            fecha_fin = hoy if anio == hoy.year else date(anio, 12, 31)
        return fecha_inicio, fecha_fin

    # Ventas de huevos (diario)
    fecha_inicio, fecha_fin = rango_diario()
    ventas_huevos_diarias_raw = db.session.query(
        func.date(Venta.fecha_venta).label('fecha'),
        func.sum(DetalleVenta.cantidad_huevos).label('huevos'),
        func.sum(DetalleVenta.cantidad_paneles).label('paneles'),
        func.sum(DetalleVenta.subtotal).label('total')
    ).join(DetalleVenta).filter(
        Venta.estado != 'cancelada',
        Venta.fecha_venta >= fecha_inicio,
        Venta.fecha_venta <= fecha_fin
    ).group_by(
        func.date(Venta.fecha_venta)
    ).order_by(
        func.date(Venta.fecha_venta)
    ).all()

    ventas_huevos_diarias = [
        {
            'fecha': item.fecha.strftime('%d/%m/%Y'),
            'huevos': int(item.huevos or 0),
            'paneles': int(item.paneles or 0),
            'total': float(item.total or 0),
        } for item in ventas_huevos_diarias_raw
    ]

    # Ventas de huevos (mensual)
    ventas_huevos_mensuales_raw = db.session.query(
        extract('month', Venta.fecha_venta).label('mes'),
        func.sum(DetalleVenta.cantidad_huevos).label('huevos'),
        func.sum(DetalleVenta.cantidad_paneles).label('paneles'),
        func.sum(DetalleVenta.subtotal).label('total')
    ).join(DetalleVenta).filter(
        extract('year', Venta.fecha_venta) == anio,
        Venta.estado != 'cancelada'
    ).group_by(
        extract('month', Venta.fecha_venta)
    ).order_by(
        extract('month', Venta.fecha_venta)
    ).all()

    ventas_huevos_mensuales = [
        {
            'mes': meses_nombres[int(item.mes) - 1],
            'huevos': int(item.huevos or 0),
            'paneles': int(item.paneles or 0),
            'total': float(item.total or 0),
        } for item in ventas_huevos_mensuales_raw
    ]

    # Produccion de huevos (diario)
    produccion_diaria_raw = db.session.query(
        func.date(LoteRecoleccion.fecha_recoleccion).label('fecha'),
        func.count(Huevo.id).label('huevos')
    ).join(Huevo).filter(
        LoteRecoleccion.fecha_recoleccion >= fecha_inicio,
        LoteRecoleccion.fecha_recoleccion <= fecha_fin,
        Huevo.roto == False
    ).group_by(
        func.date(LoteRecoleccion.fecha_recoleccion)
    ).order_by(
        func.date(LoteRecoleccion.fecha_recoleccion)
    ).all()

    produccion_diaria = [
        {
            'fecha': item.fecha.strftime('%d/%m/%Y'),
            'huevos': int(item.huevos or 0),
            'paneles': int((item.huevos or 0) // 30),
            'sobrantes': int((item.huevos or 0) % 30),
        } for item in produccion_diaria_raw
    ]

    # Produccion de huevos (mensual)
    produccion_mensual_raw = db.session.query(
        extract('month', LoteRecoleccion.fecha_recoleccion).label('mes'),
        func.count(Huevo.id).label('huevos')
    ).join(Huevo).filter(
        extract('year', LoteRecoleccion.fecha_recoleccion) == anio,
        Huevo.roto == False
    ).group_by(
        extract('month', LoteRecoleccion.fecha_recoleccion)
    ).order_by(
        extract('month', LoteRecoleccion.fecha_recoleccion)
    ).all()

    produccion_mensual = [
        {
            'mes': meses_nombres[int(item.mes) - 1],
            'huevos': int(item.huevos or 0),
            'paneles': int((item.huevos or 0) // 30),
            'sobrantes': int((item.huevos or 0) % 30),
        } for item in produccion_mensual_raw
    ]

    # Disponibilidad actual
    disponibilidad_huevos_raw = db.session.query(
        CategoriaHuevo.nombre,
        CategoriaHuevo.precio_venta,
        func.count(Huevo.id).label('cantidad')
    ).join(
        Huevo, CategoriaHuevo.id == Huevo.categoria_id
    ).filter(
        Huevo.roto == False,
        Huevo.vendido == False
    ).group_by(
        CategoriaHuevo.id, CategoriaHuevo.nombre, CategoriaHuevo.precio_venta
    ).order_by(
        CategoriaHuevo.peso_min
    ).all()

    disponibilidad_huevos = [
        {
            'nombre': item.nombre,
            'huevos': int(item.cantidad or 0),
            'paneles': int((item.cantidad or 0) // 30),
            'sobrantes': int((item.cantidad or 0) % 30),
            'precio': float(item.precio_venta or 0)
        } for item in disponibilidad_huevos_raw
    ]

    # Mortalidad mensual
    mortalidad_mensual = [0] * 12
    mortalidad_raw = db.session.query(
        extract('month', RegistroMortalidad.fecha_registro).label('mes'),
        func.sum(RegistroMortalidad.cantidad).label('cantidad')
    ).filter(
        extract('year', RegistroMortalidad.fecha_registro) == anio
    ).group_by(
        extract('month', RegistroMortalidad.fecha_registro)
    ).order_by(
        extract('month', RegistroMortalidad.fecha_registro)
    ).all()
    for item in mortalidad_raw:
        if item.mes:
            mortalidad_mensual[int(item.mes) - 1] = int(item.cantidad or 0)

    # Ventas de gallinas (filtrado)
    ventas_gallinas_query = VentaGallinas.query
    if periodo == 'diario':
        ventas_gallinas_query = ventas_gallinas_query.filter(
            VentaGallinas.fecha_venta >= fecha_inicio,
            VentaGallinas.fecha_venta <= fecha_fin
        )
    else:
        ventas_gallinas_query = ventas_gallinas_query.filter(
            extract('year', VentaGallinas.fecha_venta) == anio
        )

    ventas_gallinas = [
        {
            'fecha': row.fecha_venta.strftime('%d/%m/%Y'),
            'cantidad': int(row.cantidad or 0),
            'precio_unitario': float(row.precio_unitario or 0),
            'precio_total': float(row.precio_total or 0),
            'comprador': row.comprador or ''
        } for row in ventas_gallinas_query.order_by(VentaGallinas.fecha_venta.desc()).all()
    ]

    return {
        'ventas_huevos_diarias': ventas_huevos_diarias,
        'ventas_huevos_mensuales': ventas_huevos_mensuales,
        'produccion_diaria': produccion_diaria,
        'produccion_mensual': produccion_mensual,
        'disponibilidad_huevos': disponibilidad_huevos,
        'mortalidad': {
            'labels': meses_nombres,
            'data': mortalidad_mensual
        },
        'ventas_gallinas': ventas_gallinas
    }


@bp.route('/api/dashboard/export-excel', methods=['POST'])
@login_required
def export_dashboard_excel():
    """Exporta reportes con graficas en Excel"""
    payload = request.get_json(silent=True) or {}
    periodo = payload.get('periodo', 'diario')
    anio = payload.get('anio')
    mes = payload.get('mes')
    charts = payload.get('charts', {}) or {}

    hoy = datetime.now().date()
    if not anio:
        anio = hoy.year

    data = _build_reportes_data(periodo, anio, mes)

    wb = Workbook()
    wb.remove(wb.active)

    def add_sheet_with_table(title, headers, rows):
        ws = wb.create_sheet(title)
        ws.append(headers)
        for row in rows:
            ws.append(row)
        return ws

    add_sheet_with_table(
        'Ventas Huevos Diario',
        ['Fecha', 'Huevos', 'Paneles', 'Total'],
        [[r['fecha'], r['huevos'], r['paneles'], r['total']] for r in data['ventas_huevos_diarias']]
    )
    add_sheet_with_table(
        'Ventas Huevos Mensual',
        ['Mes', 'Huevos', 'Paneles', 'Total'],
        [[r['mes'], r['huevos'], r['paneles'], r['total']] for r in data['ventas_huevos_mensuales']]
    )
    add_sheet_with_table(
        'Produccion Diario',
        ['Fecha', 'Huevos', 'Paneles', 'Sobrantes'],
        [[r['fecha'], r['huevos'], r['paneles'], r['sobrantes']] for r in data['produccion_diaria']]
    )
    add_sheet_with_table(
        'Produccion Mensual',
        ['Mes', 'Huevos', 'Paneles', 'Sobrantes'],
        [[r['mes'], r['huevos'], r['paneles'], r['sobrantes']] for r in data['produccion_mensual']]
    )
    add_sheet_with_table(
        'Disponibilidad',
        ['Categoria', 'Huevos', 'Paneles', 'Sobrantes', 'Precio'],
        [[r['nombre'], r['huevos'], r['paneles'], r['sobrantes'], r['precio']] for r in data['disponibilidad_huevos']]
    )
    add_sheet_with_table(
        'Mortalidad',
        ['Mes', 'Cantidad'],
        [[label, cantidad] for label, cantidad in zip(data['mortalidad']['labels'], data['mortalidad']['data'])]
    )
    add_sheet_with_table(
        'Ventas Gallinas',
        ['Fecha', 'Cantidad', 'Precio Unitario', 'Precio Total', 'Comprador'],
        [[r['fecha'], r['cantidad'], r['precio_unitario'], r['precio_total'], r['comprador']] for r in data['ventas_gallinas']]
    )

    # Graficas
    ws_img = wb.create_sheet('Graficas')
    row_cursor = 1
    for key, data_url in charts.items():
        if not data_url or not data_url.startswith('data:image'):
            continue
        header = key.replace('_', ' ').title()
        ws_img.cell(row=row_cursor, column=1, value=header)
        row_cursor += 1
        b64_data = data_url.split(',', 1)[1]
        image_bytes = base64.b64decode(b64_data)
        image = PILImage.open(io.BytesIO(image_bytes))
        img = XLImage(image)
        img.anchor = f'A{row_cursor}'
        ws_img.add_image(img)
        row_cursor += 20

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    response = make_response(output.read())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename=reportes_dashboard_{periodo}_{anio}.xlsx'
    return response
