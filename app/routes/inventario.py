from flask import Blueprint, render_template, redirect, url_for, flash, request, jsonify
from flask_login import login_required, current_user
from flask_wtf.csrf import generate_csrf
from app import db, csrf
from app.models import CategoriaHuevo, LoteRecoleccion, Huevo, InventarioHuevos, User, Pesa, Gasto, Venta, Cliente, VentaGallinas, LoteGallinas
from app.utils.excel import create_excel_response
from app.utils.timezone import now_colombia, today_colombia
from datetime import datetime, date, time
from openpyxl import load_workbook
import os
from werkzeug.utils import secure_filename
import secrets

inventario_bp = Blueprint('inventario', __name__, url_prefix='/inventario')

# =============================================================================
# DASHBOARD DE INVENTARIO
# =============================================================================

@inventario_bp.route('/')
@inventario_bp.route('/dashboard')
@login_required
def dashboard():
    """Dashboard principal del inventario"""
    per_page = 10

    def build_compact_pages(current_page, total_pages, edge=2, window=2):
        if total_pages <= 1:
            return [1]

        pages = set()
        pages.update(range(1, min(edge, total_pages) + 1))
        pages.update(range(max(1, current_page - window), min(total_pages, current_page + window) + 1))
        pages.update(range(max(1, total_pages - edge + 1), total_pages + 1))

        ordered = sorted(pages)
        compact = []
        prev = None
        for page in ordered:
            if prev is not None and page - prev > 1:
                compact.append(None)
            compact.append(page)
            prev = page
        return compact

    rotos_q = (request.args.get('rotos_q') or '').strip()
    rotos_fecha_raw = (request.args.get('rotos_fecha') or '').strip()
    rotos_page = max(request.args.get('rotos_page', 1, type=int) or 1, 1)

    lotes_q = (request.args.get('lotes_q') or '').strip()
    lotes_fecha_raw = (request.args.get('lotes_fecha') or '').strip()
    lotes_page = max(request.args.get('lotes_page', 1, type=int) or 1, 1)
    # EstadÃ­sticas generales
    stats = {
        'lotes_hoy': LoteRecoleccion.query.filter_by(fecha_recoleccion=date.today()).count(),
        'lotes_activos': LoteRecoleccion.query.filter_by(estado='EN_PROCESO').count(),
        'total_huevos_hoy': db.session.query(db.func.count(Huevo.id)).join(LoteRecoleccion).filter(
            LoteRecoleccion.fecha_recoleccion == date.today(),
            Huevo.roto == False
        ).scalar() or 0,
        'huevos_rotos_hoy': db.session.query(db.func.count(Huevo.id)).join(LoteRecoleccion).filter(
            LoteRecoleccion.fecha_recoleccion == date.today(),
            Huevo.roto == True
        ).scalar() or 0
    }
    
    # Inventario por categorÃ­as
    inventario = db.session.query(
        CategoriaHuevo.nombre,
        CategoriaHuevo.peso_min,
        CategoriaHuevo.peso_max,
        db.func.count(Huevo.id).label('cantidad')
    ).outerjoin(Huevo, 
        db.and_(Huevo.categoria_id == CategoriaHuevo.id, Huevo.roto == False, Huevo.vendido == False)
    ).filter(CategoriaHuevo.activo == True).group_by(CategoriaHuevo.id).all()
    
    # Huevos rotos acumulados hasta hoy (con filtros + paginacion)
    fin_hoy = datetime.combine(date.today(), time.max)
    query_huevos_rotos = (
        db.session.query(
            Huevo.id.label('huevo_id'),
            Huevo.timestamp.label('timestamp'),
            Huevo.peso.label('peso'),
            LoteRecoleccion.id.label('lote_id'),
            LoteRecoleccion.numero_lote.label('numero_lote'),
            CategoriaHuevo.nombre.label('categoria_nombre'),
            User.first_name.label('usuario_first_name'),
            User.last_name.label('usuario_last_name'),
            User.username.label('usuario_username')
        )
        .join(LoteRecoleccion, Huevo.lote_id == LoteRecoleccion.id)
        .outerjoin(CategoriaHuevo, Huevo.categoria_id == CategoriaHuevo.id)
        .outerjoin(User, LoteRecoleccion.usuario_id == User.id)
        .filter(
            Huevo.roto == True,
            Huevo.timestamp <= fin_hoy
        )
    )

    if rotos_q:
        pattern = f'%{rotos_q}%'
        query_huevos_rotos = query_huevos_rotos.filter(LoteRecoleccion.numero_lote.ilike(pattern))

    if rotos_fecha_raw:
        try:
            rotos_fecha = datetime.strptime(rotos_fecha_raw, '%Y-%m-%d').date()
            query_huevos_rotos = query_huevos_rotos.filter(db.func.date(Huevo.timestamp) == rotos_fecha)
        except ValueError:
            rotos_fecha_raw = ''

    total_huevos_rotos = query_huevos_rotos.count()
    rotos_pages = (total_huevos_rotos + per_page - 1) // per_page if total_huevos_rotos > 0 else 1
    if rotos_page > rotos_pages:
        rotos_page = rotos_pages
    rotos_pagination_items = build_compact_pages(rotos_page, rotos_pages)

    huevos_rotos_hasta_fecha = (
        query_huevos_rotos
        .order_by(Huevo.timestamp.desc())
        .offset((rotos_page - 1) * per_page)
        .limit(per_page)
        .all()
    )

    # Lotes recientes (con filtros + paginacion)
    query_lotes = LoteRecoleccion.query

    if lotes_q:
        pattern = f'%{lotes_q}%'
        query_lotes = query_lotes.filter(LoteRecoleccion.numero_lote.ilike(pattern))

    if lotes_fecha_raw:
        try:
            lotes_fecha = datetime.strptime(lotes_fecha_raw, '%Y-%m-%d').date()
            query_lotes = query_lotes.filter(LoteRecoleccion.fecha_recoleccion == lotes_fecha)
        except ValueError:
            lotes_fecha_raw = ''

    total_lotes = query_lotes.count()
    lotes_pages = (total_lotes + per_page - 1) // per_page if total_lotes > 0 else 1
    if lotes_page > lotes_pages:
        lotes_page = lotes_pages
    lotes_pagination_items = build_compact_pages(lotes_page, lotes_pages)

    lotes_recientes = (
        query_lotes
        .order_by(LoteRecoleccion.created_at.desc())
        .offset((lotes_page - 1) * per_page)
        .limit(per_page)
        .all()
    )
    
    return render_template('inventario/dashboard.html',
                         stats=stats,
                         inventario=inventario,
                         lotes_recientes=lotes_recientes,
                         huevos_rotos_hasta_fecha=huevos_rotos_hasta_fecha,
                         total_huevos_rotos=total_huevos_rotos,
                         rotos_q=rotos_q,
                         rotos_fecha=rotos_fecha_raw,
                         rotos_page=rotos_page,
                         rotos_pages=rotos_pages,
                         rotos_pagination_items=rotos_pagination_items,
                         total_lotes=total_lotes,
                         lotes_q=lotes_q,
                         lotes_fecha=lotes_fecha_raw,
                         lotes_page=lotes_page,
                         lotes_pages=lotes_pages,
                         lotes_pagination_items=lotes_pagination_items)

# =============================================================================
# GESTIÃ“N DE CATEGORÃAS
# =============================================================================

@inventario_bp.route('/categorias')
@login_required
def categorias():
    """Vista de gestiÃ³n de categorÃ­as de huevos"""
    categorias = CategoriaHuevo.query.order_by(CategoriaHuevo.peso_min.asc()).all()
    return render_template('inventario/categorias.html', 
                         categorias=categorias,
                         csrf_token=generate_csrf)

@inventario_bp.route('/categorias/crear', methods=['POST'])
@login_required
def crear_categoria():
    """Crear nueva categorÃ­a de huevos"""
    if not current_user.is_admin:
        flash('No tienes permisos para crear categorÃ­as', 'error')
        return redirect(url_for('inventario.categorias'))
    
    try:
        categoria = CategoriaHuevo(
            nombre=request.form.get('nombre'),
            peso_min=float(request.form.get('peso_min')),
            peso_max=float(request.form.get('peso_max')),
            precio_venta=float(request.form.get('precio_venta', 0))
        )
        
        db.session.add(categoria)
        db.session.commit()
        
        flash(f'CategorÃ­a {categoria.nombre} creada correctamente', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error al crear la categorÃ­a: {str(e)}', 'error')
    
    return redirect(url_for('inventario.categorias'))

@inventario_bp.route('/categorias/<int:categoria_id>/editar', methods=['GET'])
@login_required
def obtener_categoria(categoria_id):
    """Obtener datos de una categorÃ­a para ediciÃ³n"""
    if not current_user.is_admin:
        return jsonify({'success': False, 'message': 'Sin permisos'})
    
    try:
        categoria = CategoriaHuevo.query.get_or_404(categoria_id)
        return jsonify({
            'success': True,
            'categoria': {
                'id': categoria.id,
                'nombre': categoria.nombre,
                'peso_min': categoria.peso_min,
                'peso_max': categoria.peso_max,
                'precio_venta': float(categoria.precio_venta),
                'activo': categoria.activo
            }
        })
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})

@inventario_bp.route('/categorias/<int:categoria_id>/actualizar', methods=['POST'])
@login_required
def actualizar_categoria(categoria_id):
    """Actualizar una categorÃ­a existente"""
    if not current_user.is_admin:
        return jsonify({'success': False, 'message': 'Sin permisos'})
    
    try:
        categoria = CategoriaHuevo.query.get_or_404(categoria_id)
        
        categoria.nombre = request.form.get('nombre')
        categoria.peso_min = float(request.form.get('peso_min'))
        categoria.peso_max = float(request.form.get('peso_max'))
        categoria.precio_venta = float(request.form.get('precio_venta', 0))
        
        db.session.commit()
        
        flash(f'CategorÃ­a {categoria.nombre} actualizada correctamente', 'success')
        return jsonify({'success': True})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)})

@inventario_bp.route('/categorias/<int:categoria_id>/cambiar-estado', methods=['POST'])
@login_required
def cambiar_estado_categoria(categoria_id):
    """Cambiar el estado activo/inactivo de una categorÃ­a"""
    if not current_user.is_admin:
        return jsonify({'success': False, 'message': 'Sin permisos'})
    
    try:
        categoria = CategoriaHuevo.query.get_or_404(categoria_id)
        nuevo_estado = request.json.get('activo', True)
        categoria.activo = nuevo_estado
        
        db.session.commit()
        
        estado_texto = 'activada' if nuevo_estado else 'desactivada'
        flash(f'CategorÃ­a {categoria.nombre} {estado_texto} correctamente', 'success')
        
        return jsonify({
            'success': True,
            'message': f'CategorÃ­a {estado_texto} correctamente'
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)})

# =============================================================================
# GESTIÃ“N DE LOTES DE RECOLECCIÃ“N
# =============================================================================

# =============================================================================
# GESTION DE PESAS
# =============================================================================

@inventario_bp.route('/pesas')
@login_required
def pesas():
    """Vista de gestion de pesas"""
    if not current_user.is_admin:
        flash('No tienes permisos para gestionar pesas', 'error')
        return redirect(url_for('inventario.dashboard'))

    pesas = Pesa.query.order_by(Pesa.id.asc()).all()
    return render_template('inventario/pesas.html', pesas=pesas, csrf_token=generate_csrf)


@inventario_bp.route('/pesas/crear', methods=['POST'])
@login_required
def crear_pesa():
    """Crear nueva pesa"""
    if not current_user.is_admin:
        flash('No tienes permisos para crear pesas', 'error')
        return redirect(url_for('inventario.pesas'))

    try:
        pesa = Pesa(
            nombre=request.form.get('nombre'),
            base_url=request.form.get('base_url'),
            token_api=request.form.get('token_api'),
            puerto=request.form.get('puerto'),
            baud=int(request.form.get('baud', 9600)),
            tolerancia=float(request.form.get('tolerancia', 1.0)),
            reset_threshold=float(request.form.get('reset_threshold', 1.0)),
            activo=True
        )
        db.session.add(pesa)
        db.session.commit()
        flash(f'Pesa #{pesa.id} creada correctamente', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error al crear la pesa: {str(e)}', 'error')

    return redirect(url_for('inventario.pesas'))


@inventario_bp.route('/pesas/<int:pesa_id>/editar', methods=['GET'])
@login_required
def obtener_pesa(pesa_id):
    """Obtener datos de una pesa para edicion"""
    if not current_user.is_admin:
        return jsonify({'success': False, 'message': 'Sin permisos'}), 403

    pesa = Pesa.query.get_or_404(pesa_id)
    return jsonify({
        'success': True,
        'pesa': {
            'id': pesa.id,
            'nombre': pesa.nombre,
            'base_url': pesa.base_url,
            'token_api': pesa.token_api,
            'puerto': pesa.puerto,
            'baud': pesa.baud,
            'tolerancia': pesa.tolerancia,
            'reset_threshold': pesa.reset_threshold,
            'activo': pesa.activo
        }
    })


@inventario_bp.route('/pesas/<int:pesa_id>/actualizar', methods=['POST'])
@login_required
def actualizar_pesa(pesa_id):
    """Actualizar configuracion de una pesa"""
    if not current_user.is_admin:
        return jsonify({'success': False, 'message': 'Sin permisos'}), 403

    try:
        pesa = Pesa.query.get_or_404(pesa_id)
        pesa.nombre = request.form.get('nombre')
        pesa.base_url = request.form.get('base_url')
        pesa.token_api = request.form.get('token_api')
        pesa.puerto = request.form.get('puerto')
        pesa.baud = int(request.form.get('baud', 9600))
        pesa.tolerancia = float(request.form.get('tolerancia', 1.0))
        pesa.reset_threshold = float(request.form.get('reset_threshold', 1.0))

        db.session.commit()
        flash(f'Pesa #{pesa.id} actualizada correctamente', 'success')
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)})


@inventario_bp.route('/pesas/<int:pesa_id>/cambiar-estado', methods=['POST'])
@login_required
def cambiar_estado_pesa(pesa_id):
    """Activar o desactivar una pesa"""
    if not current_user.is_admin:
        return jsonify({'success': False, 'message': 'Sin permisos'}), 403

    try:
        pesa = Pesa.query.get_or_404(pesa_id)
        pesa.activo = bool(request.json.get('activo', True))
        db.session.commit()
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)})


# =============================================================================
# GESTION DE GASTOS
# =============================================================================

def _build_gastos_query():
    query = Gasto.query.join(User, Gasto.usuario_id == User.id)

    fecha_desde = request.args.get('fecha_desde', '').strip()
    fecha_hasta = request.args.get('fecha_hasta', '').strip()
    tipo = request.args.get('tipo', '').strip().lower()
    usuario_id = request.args.get('usuario_id', type=int)
    q = request.args.get('q', '').strip()

    if fecha_desde:
        try:
            d = datetime.strptime(fecha_desde, '%Y-%m-%d').date()
            query = query.filter(Gasto.fecha_hora >= datetime.combine(d, time.min))
        except ValueError:
            pass

    if fecha_hasta:
        try:
            d = datetime.strptime(fecha_hasta, '%Y-%m-%d').date()
            query = query.filter(Gasto.fecha_hora <= datetime.combine(d, time.max))
        except ValueError:
            pass

    if tipo in {'insumos', 'servicios', 'otros'}:
        query = query.filter(Gasto.tipo == tipo)

    if usuario_id:
        query = query.filter(Gasto.usuario_id == usuario_id)

    if q:
        pattern = f"%{q}%"
        query = query.filter(
            db.or_(
                Gasto.descripcion.ilike(pattern),
                User.first_name.ilike(pattern),
                User.last_name.ilike(pattern),
                User.username.ilike(pattern),
            )
        )

    return query


@inventario_bp.route('/gastos')
@login_required
def gastos():
    """Vista para gestionar gastos con filtros"""
    if not current_user.is_admin:
        flash('No tienes permisos para gestionar gastos', 'error')
        return redirect(url_for('inventario.dashboard'))

    page = request.args.get('page', 1, type=int)
    query = _build_gastos_query().order_by(Gasto.fecha_hora.desc(), Gasto.id.desc())
    gastos_paginados = query.paginate(page=page, per_page=30, error_out=False)

    total_filtrado = query.with_entities(db.func.coalesce(db.func.sum(Gasto.valor), 0)).scalar() or 0
    usuarios = User.query.filter_by(is_active=True).order_by(User.first_name.asc(), User.last_name.asc()).all()

    return render_template(
        'inventario/gastos.html',
        gastos=gastos_paginados,
        usuarios=usuarios,
        total_filtrado=float(total_filtrado),
        today_str=date.today().strftime('%Y-%m-%d'),
        now_str=datetime.now().strftime('%H:%M'),
        csrf_token=generate_csrf,
    )


@inventario_bp.route('/gastos/crear', methods=['POST'])
@login_required
def crear_gasto():
    """Crear un gasto"""
    if not current_user.is_admin:
        flash('No tienes permisos para crear gastos', 'error')
        return redirect(url_for('inventario.gastos'))

    try:
        tipo = (request.form.get('tipo') or '').strip().lower()
        if tipo not in {'insumos', 'servicios', 'otros'}:
            raise ValueError('Tipo de gasto invalido')

        valor = float(request.form.get('valor', 0))
        if valor <= 0:
            raise ValueError('El valor debe ser mayor que cero')

        descripcion = (request.form.get('descripcion') or '').strip()
        if not descripcion:
            raise ValueError('La descripcion es obligatoria')

        fecha_str = (request.form.get('fecha') or '').strip()
        hora_str = (request.form.get('hora') or '').strip()
        if fecha_str:
            fecha = datetime.strptime(fecha_str, '%Y-%m-%d').date()
        else:
            fecha = date.today()
        if hora_str:
            hora = datetime.strptime(hora_str, '%H:%M').time()
        else:
            hora = datetime.now().time().replace(second=0, microsecond=0)

        gasto = Gasto(
            fecha_hora=datetime.combine(fecha, hora),
            valor=valor,
            tipo=tipo,
            descripcion=descripcion,
            usuario_id=current_user.id,
        )
        db.session.add(gasto)
        db.session.commit()
        flash('Gasto creado correctamente', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error al crear gasto: {str(e)}', 'error')

    return redirect(url_for('inventario.gastos'))


@inventario_bp.route('/gastos/<int:gasto_id>/actualizar', methods=['POST'])
@login_required
def actualizar_gasto(gasto_id):
    """Actualizar un gasto"""
    if not current_user.is_admin:
        flash('No tienes permisos para editar gastos', 'error')
        return redirect(url_for('inventario.gastos'))

    gasto = Gasto.query.get_or_404(gasto_id)

    try:
        tipo = (request.form.get('tipo') or '').strip().lower()
        if tipo not in {'insumos', 'servicios', 'otros'}:
            raise ValueError('Tipo de gasto invalido')

        valor = float(request.form.get('valor', 0))
        if valor <= 0:
            raise ValueError('El valor debe ser mayor que cero')

        descripcion = (request.form.get('descripcion') or '').strip()
        if not descripcion:
            raise ValueError('La descripcion es obligatoria')

        fecha_str = (request.form.get('fecha') or '').strip()
        hora_str = (request.form.get('hora') or '').strip()
        if not fecha_str or not hora_str:
            raise ValueError('La fecha y la hora son obligatorias')

        fecha = datetime.strptime(fecha_str, '%Y-%m-%d').date()
        hora = datetime.strptime(hora_str, '%H:%M').time()

        gasto.fecha_hora = datetime.combine(fecha, hora)
        gasto.tipo = tipo
        gasto.valor = valor
        gasto.descripcion = descripcion
        gasto.usuario_id = current_user.id
        db.session.commit()
        flash('Gasto actualizado correctamente', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error al actualizar gasto: {str(e)}', 'error')

    return redirect(url_for('inventario.gastos'))


@inventario_bp.route('/gastos/<int:gasto_id>/eliminar', methods=['POST'])
@login_required
def eliminar_gasto(gasto_id):
    """Eliminar un gasto"""
    if not current_user.is_admin:
        flash('No tienes permisos para eliminar gastos', 'error')
        return redirect(url_for('inventario.gastos'))

    gasto = Gasto.query.get_or_404(gasto_id)

    try:
        db.session.delete(gasto)
        db.session.commit()
        flash('Gasto eliminado correctamente', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error al eliminar gasto: {str(e)}', 'error')

    return redirect(url_for('inventario.gastos'))


@inventario_bp.route('/gastos/export/excel')
@login_required
def exportar_gastos_excel():
    """Exportar gastos filtrados a Excel"""
    if not current_user.is_admin:
        flash('No tienes permisos para exportar gastos', 'error')
        return redirect(url_for('inventario.gastos'))

    rows_query = (
        _build_gastos_query()
        .order_by(Gasto.fecha_hora.desc(), Gasto.id.desc())
        .with_entities(
            Gasto.fecha_hora,
            Gasto.valor,
            Gasto.tipo,
            Gasto.descripcion,
            User.username,
            User.first_name,
            User.last_name,
        )
        .all()
    )

    rows = [
        [
            item.fecha_hora.date(),
            item.fecha_hora.strftime('%H:%M'),
            float(item.valor or 0),
            item.tipo.capitalize(),
            item.descripcion,
            item.username,
            f'{item.first_name} {item.last_name}'.strip(),
        ]
        for item in rows_query
    ]

    return create_excel_response(
        'inventario_gastos.xlsx',
        'Gastos',
        ['Fecha', 'Hora', 'Valor', 'Tipo', 'Descripcion', 'Usuario', 'Nombre'],
        rows,
    )


def _build_movimientos_combinados():
    fecha_desde = request.args.get('fecha_desde', '').strip()
    fecha_hasta = request.args.get('fecha_hasta', '').strip()
    tipo_mov = request.args.get('tipo_mov', '').strip().lower()  # venta | gasto
    tipo_gasto = request.args.get('tipo_gasto', '').strip().lower()
    estado_venta = request.args.get('estado_venta', '').strip().lower()
    q = request.args.get('q', '').strip()

    dt_desde = None
    dt_hasta = None
    if fecha_desde:
        try:
            d = datetime.strptime(fecha_desde, '%Y-%m-%d').date()
            dt_desde = datetime.combine(d, time.min)
        except ValueError:
            pass
    if fecha_hasta:
        try:
            d = datetime.strptime(fecha_hasta, '%Y-%m-%d').date()
            dt_hasta = datetime.combine(d, time.max)
        except ValueError:
            pass

    movimientos = []

    # Ventas (ingresos)
    if tipo_mov in {'', 'venta'}:
        ventas_q = Venta.query.outerjoin(Cliente, Venta.cliente_id == Cliente.id).join(User, Venta.vendedor_id == User.id)
        if dt_desde:
            ventas_q = ventas_q.filter(Venta.fecha_venta >= dt_desde)
        if dt_hasta:
            ventas_q = ventas_q.filter(Venta.fecha_venta <= dt_hasta)
        if estado_venta:
            ventas_q = ventas_q.filter(Venta.estado == estado_venta)
        if q:
            pattern = f'%{q}%'
            ventas_q = ventas_q.filter(
                db.or_(
                    Venta.numero_venta.ilike(pattern),
                    Venta.observaciones.ilike(pattern),
                    User.username.ilike(pattern),
                    User.first_name.ilike(pattern),
                    User.last_name.ilike(pattern),
                    Cliente.nombre.ilike(pattern),
                    Cliente.apellido.ilike(pattern),
                )
            )

        for v in ventas_q.order_by(Venta.fecha_venta.desc(), Venta.id.desc()).all():
            cliente = v.cliente.get_nombre_completo() if v.cliente else 'Cliente contado'
            movimientos.append({
                'fecha_hora': v.fecha_venta,
                'movimiento': 'VENTA',
                'subtipo': v.tipo_pago.upper(),
                'referencia': v.numero_venta,
                'descripcion': f'Venta a {cliente}',
                'usuario': v.vendedor.get_full_name() if v.vendedor else '',
                'estado': v.estado,
                'ingreso': float(v.total or 0),
                'egreso': 0.0,
            })

        # Ventas de gallinas (tambien ingresos)
        ventas_gallinas_q = VentaGallinas.query\
            .join(User, VentaGallinas.usuario_id == User.id)\
            .join(LoteGallinas, VentaGallinas.lote_gallinas_id == LoteGallinas.id)
        if dt_desde:
            ventas_gallinas_q = ventas_gallinas_q.filter(VentaGallinas.fecha_venta >= dt_desde.date())
        if dt_hasta:
            ventas_gallinas_q = ventas_gallinas_q.filter(VentaGallinas.fecha_venta <= dt_hasta.date())
        if estado_venta and estado_venta != 'completada':
            ventas_gallinas_q = ventas_gallinas_q.filter(db.text('1=0'))
        if q:
            pattern = f'%{q}%'
            ventas_gallinas_q = ventas_gallinas_q.filter(
                db.or_(
                    VentaGallinas.comprador.ilike(pattern),
                    VentaGallinas.observaciones.ilike(pattern),
                    LoteGallinas.numero_lote.ilike(pattern),
                    User.username.ilike(pattern),
                    User.first_name.ilike(pattern),
                    User.last_name.ilike(pattern),
                )
            )

        for vg in ventas_gallinas_q.order_by(VentaGallinas.fecha_venta.desc(), VentaGallinas.id.desc()).all():
            fecha_hora = datetime.combine(vg.fecha_venta, time.min) if vg.fecha_venta else None
            comprador = vg.comprador if vg.comprador else 'Cliente contado'
            movimientos.append({
                'fecha_hora': fecha_hora,
                'movimiento': 'VENTA',
                'subtipo': 'GALLINAS',
                'referencia': f'VG-{vg.id}',
                'descripcion': f'Venta de {vg.cantidad} gallinas del lote {vg.lote_gallinas.numero_lote} a {comprador}',
                'usuario': vg.usuario.get_full_name() if vg.usuario else '',
                'estado': 'completada',
                'ingreso': float(vg.precio_total or 0),
                'egreso': 0.0,
            })

    # Gastos (egresos)
    if tipo_mov in {'', 'gasto'}:
        gastos_q = Gasto.query.join(User, Gasto.usuario_id == User.id)
        if dt_desde:
            gastos_q = gastos_q.filter(Gasto.fecha_hora >= dt_desde)
        if dt_hasta:
            gastos_q = gastos_q.filter(Gasto.fecha_hora <= dt_hasta)
        if tipo_gasto in {'insumos', 'servicios', 'otros'}:
            gastos_q = gastos_q.filter(Gasto.tipo == tipo_gasto)
        if q:
            pattern = f'%{q}%'
            gastos_q = gastos_q.filter(
                db.or_(
                    Gasto.descripcion.ilike(pattern),
                    User.username.ilike(pattern),
                    User.first_name.ilike(pattern),
                    User.last_name.ilike(pattern),
                )
            )

        for g in gastos_q.order_by(Gasto.fecha_hora.desc(), Gasto.id.desc()).all():
            movimientos.append({
                'fecha_hora': g.fecha_hora,
                'movimiento': 'GASTO',
                'subtipo': g.tipo.upper(),
                'referencia': f'G-{g.id}',
                'descripcion': g.descripcion,
                'usuario': g.usuario.get_full_name() if g.usuario else '',
                'estado': '-',
                'ingreso': 0.0,
                'egreso': float(g.valor or 0),
            })

    movimientos.sort(key=lambda x: (x['fecha_hora'], x['referencia']), reverse=True)

    total_ingresos = sum(m['ingreso'] for m in movimientos)
    total_egresos = sum(m['egreso'] for m in movimientos)
    balance = total_ingresos - total_egresos

    return movimientos, total_ingresos, total_egresos, balance


@inventario_bp.route('/movimientos')
@login_required
def movimientos_combinados():
    """Tabla combinada de ventas y gastos"""
    if not current_user.is_admin:
        flash('No tienes permisos para ver movimientos combinados', 'error')
        return redirect(url_for('inventario.dashboard'))

    movimientos, total_ingresos, total_egresos, balance = _build_movimientos_combinados()

    page = request.args.get('page', 1, type=int)
    per_page = 40
    total_items = len(movimientos)
    total_pages = max((total_items + per_page - 1) // per_page, 1)
    if page < 1:
        page = 1
    if page > total_pages:
        page = total_pages
    start = (page - 1) * per_page
    end = start + per_page
    items = movimientos[start:end]

    return render_template(
        'inventario/movimientos.html',
        movimientos=items,
        page=page,
        total_pages=total_pages,
        total_items=total_items,
        total_ingresos=float(total_ingresos),
        total_egresos=float(total_egresos),
        balance=float(balance),
    )


@inventario_bp.route('/movimientos/export/excel')
@login_required
def exportar_movimientos_excel():
    """Exportar tabla combinada de ventas y gastos"""
    if not current_user.is_admin:
        flash('No tienes permisos para exportar movimientos', 'error')
        return redirect(url_for('inventario.movimientos_combinados'))

    movimientos, total_ingresos, total_egresos, balance = _build_movimientos_combinados()
    rows = [
        [
            m['fecha_hora'].date() if m['fecha_hora'] else '',
            m['fecha_hora'].strftime('%H:%M') if m['fecha_hora'] else '',
            m['movimiento'],
            m['subtipo'],
            m['referencia'],
            m['descripcion'],
            m['usuario'],
            m['estado'],
            float(m['ingreso']),
            float(m['egreso']),
        ]
        for m in movimientos
    ]
    rows.append(['', '', '', '', '', 'TOTAL INGRESOS', '', '', float(total_ingresos), 0.0])
    rows.append(['', '', '', '', '', 'TOTAL EGRESOS', '', '', 0.0, float(total_egresos)])
    rows.append(['', '', '', '', '', 'BALANCE', '', '', float(balance), 0.0])

    return create_excel_response(
        'movimientos_ventas_gastos.xlsx',
        'Movimientos',
        ['Fecha', 'Hora', 'Movimiento', 'Subtipo', 'Referencia', 'Descripcion', 'Usuario', 'Estado', 'Ingreso', 'Egreso'],
        rows,
    )


@inventario_bp.route('/lotes')
@login_required
def lotes():
    """Vista de gestiÃ³n de lotes de recolecciÃ³n"""
    page = request.args.get('page', 1, type=int)
    estado_filter = request.args.get('estado', '')
    
    query = LoteRecoleccion.query
    
    if estado_filter:
        query = query.filter_by(estado=estado_filter)
    
    lotes = query.order_by(LoteRecoleccion.created_at.desc()).paginate(
        page=page, per_page=20, error_out=False
    )
    
    return render_template('inventario/lotes.html', lotes=lotes)

@inventario_bp.route('/lotes/nuevo')
@login_required
def nuevo_lote():
    """Formulario para crear nuevo lote"""
    # Obtener lotes de gallinas activos
    from app.models import LoteGallinas
    lotes_gallinas = LoteGallinas.query.filter_by(estado='Activo').order_by(LoteGallinas.numero_lote).all()
    today = today_colombia().strftime('%Y-%m-%d')
    now_time = now_colombia().strftime('%H:%M')
    
    return render_template('inventario/nuevo_lote.html',
                         csrf_token=generate_csrf,
                         lotes_gallinas=lotes_gallinas,
                         today=today,
                         now_time=now_time)

@inventario_bp.route('/lotes/crear', methods=['POST'])
@login_required
def crear_lote():
    """Crear nuevo lote de recolecciÃ³n"""
    try:
        from app.models import LoteGallinas
        
        # Fecha/hora inmutable de Colombia para evitar lotes de otros dias.
        fecha_recoleccion_colombia = today_colombia()
        hora_inicio_colombia = now_colombia()
        
        # Obtener lote de gallinas si se seleccionÃ³
        lote_gallinas_id = request.form.get('lote_gallinas_id')
        lote_gallinas = None
        semana_produccion = None
        
        if lote_gallinas_id:
            lote_gallinas = LoteGallinas.query.get(int(lote_gallinas_id))
            if lote_gallinas:
                fecha_recoleccion = fecha_recoleccion_colombia
                
                # Si es la primera recolecciÃ³n, marcar inicio de producciÃ³n
                if not lote_gallinas.fecha_inicio_produccion:
                    lote_gallinas.fecha_inicio_produccion = fecha_recoleccion
                    semana_produccion = 1
                    flash(f'Â¡Primera recolecciÃ³n del lote {lote_gallinas.numero_lote}! Inicio de producciÃ³n registrado.', 'info')
                else:
                    # Calcular semana basada en la fecha de recolecciÃ³n vs fecha inicio producciÃ³n
                    # Sumamos 1 al resultado para que el primer dÃ­a sea dÃ­a 1 (no dÃ­a 0)
                    dias_desde_inicio = (fecha_recoleccion - lote_gallinas.fecha_inicio_produccion).days + 1
                    semana_produccion = ((dias_desde_inicio - 1) // 7) + 1  # DÃ­as 1-7=Semana1, 8-14=Semana2, etc.
        
        lote = LoteRecoleccion(
            fecha_recoleccion=fecha_recoleccion_colombia,
            hora_inicio=hora_inicio_colombia,
            usuario_id=current_user.id,
            observaciones=request.form.get('observaciones'),
            lote_gallinas_id=lote_gallinas_id if lote_gallinas_id else None,
            semana_produccion=semana_produccion
        )
        
        lote.generar_numero_lote()
        
        db.session.add(lote)
        # Si hay lote de gallinas modificado, tambiÃ©n se guardarÃ¡
        if lote_gallinas:
            db.session.add(lote_gallinas)
        
        db.session.commit()
        
        mensaje = f'Lote {lote.numero_lote} creado correctamente'
        if lote_gallinas:
            mensaje += f' - {lote_gallinas.numero_lote} (Semana {semana_produccion})'
        
        flash(mensaje, 'success')
        return redirect(url_for('inventario.pesar_huevos', lote_id=lote.id))
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error al crear el lote: {str(e)}', 'error')
        return redirect(url_for('inventario.nuevo_lote'))

# =============================================================================
# PESAJE DE HUEVOS
# =============================================================================

@inventario_bp.route('/lotes/<int:lote_id>/pesar')
@login_required
def pesar_huevos(lote_id):
    """Interfaz principal de pesaje de huevos"""
    lote = LoteRecoleccion.query.get_or_404(lote_id)
    
    # Verificar que el usuario puede acceder a este lote
    if lote.usuario_id != current_user.id and not current_user.is_admin:
        flash('No tienes permisos para acceder a este lote', 'error')
        return redirect(url_for('inventario.lotes'))
    
    # EstadÃ­sticas del lote
    huevos = Huevo.query.filter_by(lote_id=lote_id).all()
    stats_lote = {
        'total_huevos': len([h for h in huevos if not h.roto]),
        'huevos_rotos': len([h for h in huevos if h.roto]),
        'peso_total': sum([h.peso for h in huevos if not h.roto]),
        'peso_promedio': sum([h.peso for h in huevos if not h.roto]) / max(len([h for h in huevos if not h.roto]), 1)
    }
    
    # DistribuciÃ³n por categorÃ­as
    distribucion = db.session.query(
        CategoriaHuevo.nombre,
        db.func.count(Huevo.id).label('cantidad'),
        db.func.sum(Huevo.peso).label('peso_total')
    ).join(Huevo).filter(
        Huevo.lote_id == lote_id,
        Huevo.roto == False
    ).group_by(CategoriaHuevo.id, CategoriaHuevo.nombre).all()
    
    # CategorÃ­as disponibles para clasificaciÃ³n
    categorias = CategoriaHuevo.query.filter_by(activo=True).order_by(CategoriaHuevo.peso_min.asc()).all()
    pesas_activas = Pesa.query.filter_by(activo=True).order_by(Pesa.id.asc()).all()
    lotes_en_proceso_con_pesa = LoteRecoleccion.query.filter(
        LoteRecoleccion.estado == 'EN_PROCESO',
        LoteRecoleccion.pesa_id.isnot(None),
        LoteRecoleccion.id != lote_id
    ).all()
    pesas_ocupadas = {int(l.pesa_id): l.numero_lote for l in lotes_en_proceso_con_pesa if l.pesa_id}
    hay_pesa_disponible = any((p.id not in pesas_ocupadas) or (lote.pesa_id == p.id) for p in pesas_activas)
    
    return render_template('inventario/pesar_huevos.html',
                         lote=lote,
                         stats_lote=stats_lote,
                         distribucion=distribucion,
                         categorias=categorias,
                         pesas_activas=pesas_activas,
                         pesas_ocupadas=pesas_ocupadas,
                         hay_pesa_disponible=hay_pesa_disponible,
                         huevos=huevos[-50:],  # Solo los Ãºltimos 50 para mostrar
                         csrf_token=generate_csrf)

@inventario_bp.route('/lotes/<int:lote_id>/asignar-pesa', methods=['POST'])
@login_required
def asignar_pesa_lote(lote_id):
    """Asigna una pesa a un lote antes del pesaje"""
    lote = LoteRecoleccion.query.get_or_404(lote_id)

    if lote.usuario_id != current_user.id and not current_user.is_admin:
        flash('Sin permisos para modificar este lote', 'error')
        return redirect(url_for('inventario.pesar_huevos', lote_id=lote_id))

    if lote.estado != 'EN_PROCESO':
        flash('Solo puedes asignar pesa en lotes en proceso', 'error')
        return redirect(url_for('inventario.pesar_huevos', lote_id=lote_id))

    try:
        pesa_id = int(request.form.get('pesa_id', 0))
        pesa = Pesa.query.filter_by(id=pesa_id, activo=True).first()
        if not pesa:
            flash('La pesa seleccionada no es valida o esta inactiva', 'error')
            return redirect(url_for('inventario.pesar_huevos', lote_id=lote_id))
        lote_con_pesa = LoteRecoleccion.query.filter(
            LoteRecoleccion.estado == 'EN_PROCESO',
            LoteRecoleccion.pesa_id == pesa.id,
            LoteRecoleccion.id != lote_id
        ).first()
        if lote_con_pesa:
            flash(f'La pesa ID {pesa.id} ya esta en uso en el lote {lote_con_pesa.numero_lote}', 'error')
            return redirect(url_for('inventario.pesar_huevos', lote_id=lote_id))

        lote.pesa_id = pesa.id
        db.session.commit()
        flash(f'Pesa #{pesa.id} asignada correctamente al lote', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error al asignar la pesa: {str(e)}', 'error')

    return redirect(url_for('inventario.pesar_huevos', lote_id=lote_id))

@inventario_bp.route('/lotes/<int:lote_id>/pesar/manual', methods=['POST'])
@login_required
def pesar_manual(lote_id):
    """Agregar huevo manualmente"""
    lote = LoteRecoleccion.query.get_or_404(lote_id)

    if lote.usuario_id != current_user.id and not current_user.is_admin:
        return jsonify({'success': False, 'message': 'Sin permisos'})

    if not lote.pesa_id:
        return jsonify({'success': False, 'message': 'Debes asignar una pesa al lote antes de pesar'})

    try:
        payload = request.get_json(silent=True) or {}
        peso = float(payload.get('peso', 0))
        roto = bool(payload.get('roto', False))
        categoria_id_raw = payload.get('categoria_id')

        if not roto and peso <= 0:
            return jsonify({'success': False, 'message': 'Peso invalido'})

        if roto and peso <= 0:
            peso = 0

        huevo = Huevo(
            peso=peso,
            roto=roto,
            lote_id=lote_id
        )

        if roto:
            if categoria_id_raw not in (None, ''):
                try:
                    categoria_id = int(categoria_id_raw)
                except (TypeError, ValueError):
                    return jsonify({'success': False, 'message': 'Categoria invalida'})
                categoria = CategoriaHuevo.query.get(categoria_id)
                if not categoria:
                    return jsonify({'success': False, 'message': 'Categoria no encontrada'})
                huevo.categoria_id = categoria.id
        else:
            huevo.clasificar()

        db.session.add(huevo)
        lote.actualizar_estadisticas()
        db.session.commit()

        promedio = float((lote.total_peso or 0) / (lote.total_huevos or 1))
        return jsonify({
            'success': True,
            'huevo': {
                'id': huevo.id,
                'peso': huevo.peso,
                'roto': huevo.roto,
                'categoria': (
                    huevo.categoria.nombre
                    if huevo.categoria
                    else ('No disponible' if huevo.roto else 'Sin clasificar')
                )
            },
            'stats': {
                'total_huevos': int(lote.total_huevos or 0),
                'huevos_rotos': int(lote.huevos_rotos or 0),
                'peso_total': float(lote.total_peso or 0),
                'peso_promedio': promedio
            }
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)})

@inventario_bp.route('/export/excel')
@login_required
def export_excel():
    """Exportar tablas del modulo inventario a Excel"""
    tabla = request.args.get('tabla', '').strip()

    if tabla == 'lotes':
        estado_filter = request.args.get('estado', '')
        query = LoteRecoleccion.query
        if estado_filter:
            query = query.filter_by(estado=estado_filter)
        lotes = query.order_by(LoteRecoleccion.created_at.desc()).all()
        rows = [[
            l.numero_lote,
            l.fecha_recoleccion,
            l.usuario.get_full_name() if l.usuario else '',
            l.estado,
            l.total_huevos,
            l.total_peso,
            l.huevos_rotos
        ] for l in lotes]
        return create_excel_response(
            'inventario_lotes.xlsx',
            'Lotes',
            ['Numero Lote', 'Fecha Recoleccion', 'Encargado', 'Estado', 'Huevos', 'Peso Total', 'Huevos Rotos'],
            rows
        )

    if tabla == 'categorias':
        categorias = CategoriaHuevo.query.order_by(CategoriaHuevo.peso_min.asc()).all()
        rows = [[
            c.nombre,
            c.peso_min,
            c.peso_max,
            float(c.precio_venta or 0),
            'Activo' if c.activo else 'Inactivo'
        ] for c in categorias]
        return create_excel_response(
            'inventario_categorias.xlsx',
            'Categorias',
            ['Categoria', 'Peso Min', 'Peso Max', 'Precio Venta', 'Estado'],
            rows
        )

    if tabla == 'dashboard_inventario':
        inventario = db.session.query(
            CategoriaHuevo.nombre,
            CategoriaHuevo.peso_min,
            CategoriaHuevo.peso_max,
            db.func.count(Huevo.id).label('cantidad')
        ).outerjoin(
            Huevo,
            db.and_(Huevo.categoria_id == CategoriaHuevo.id, Huevo.roto == False, Huevo.vendido == False)
        ).filter(
            CategoriaHuevo.activo == True
        ).group_by(
            CategoriaHuevo.id
        ).all()
        rows = [[
            i.nombre,
            i.peso_min,
            i.peso_max,
            i.cantidad or 0,
            (i.cantidad or 0) // 30,
            (i.cantidad or 0) % 30
        ] for i in inventario]
        return create_excel_response(
            'inventario_dashboard.xlsx',
            'Inventario',
            ['Categoria', 'Peso Min', 'Peso Max', 'Cantidad', 'Panales', 'Huevos Libres'],
            rows
        )

    flash('Tipo de tabla no valido para exportar', 'error')
    return redirect(url_for('inventario.dashboard'))


@inventario_bp.route('/pesas/<int:pesa_id>/pesar/auto', methods=['POST'])
@csrf.exempt
def pesar_automatico(pesa_id):
    """Agregar huevo desde una pesa configurada por ID"""
    pesa = Pesa.query.get_or_404(pesa_id)
    if not pesa.activo:
        return jsonify({'success': False, 'message': 'Pesa inactiva'}), 409

    token_esperado = pesa.token_api or ''
    token_recibido = request.headers.get('X-Scale-Token', '')
    if not token_recibido:
        auth = request.headers.get('Authorization', '')
        if auth.lower().startswith('bearer '):
            token_recibido = auth.split(' ', 1)[1].strip()

    if not token_esperado or not token_recibido or not secrets.compare_digest(token_recibido, token_esperado):
        return jsonify({'success': False, 'message': 'Token invalido'}), 401

    lotes_activos = LoteRecoleccion.query.filter_by(pesa_id=pesa_id, estado='EN_PROCESO').all()
    if not lotes_activos:
        return jsonify({'success': False, 'message': 'No hay lote en proceso para esta pesa'}), 409
    if len(lotes_activos) > 1:
        return jsonify({'success': False, 'message': 'Hay varios lotes activos con esta pesa. Deja solo uno.'}), 409

    lote = lotes_activos[0]

    try:
        payload = request.get_json(silent=True) or {}
        peso = float(payload.get('peso', 0))
        roto = bool(payload.get('roto', False))
        categoria_id_raw = payload.get('categoria_id')

        if not roto and peso <= 0:
            return jsonify({'success': False, 'message': 'Peso invalido'}), 400

        if roto and peso <= 0:
            peso = 0

        huevo = Huevo(
            peso=peso,
            roto=roto,
            lote_id=lote.id
        )

        if roto:
            if categoria_id_raw not in (None, ''):
                try:
                    categoria_id = int(categoria_id_raw)
                except (TypeError, ValueError):
                    return jsonify({'success': False, 'message': 'Categoria invalida'}), 400
                categoria = CategoriaHuevo.query.get(categoria_id)
                if not categoria:
                    return jsonify({'success': False, 'message': 'Categoria no encontrada'}), 400
                huevo.categoria_id = categoria.id
        else:
            huevo.clasificar()

        db.session.add(huevo)
        lote.actualizar_estadisticas()
        db.session.commit()

        promedio = float((lote.total_peso or 0) / (lote.total_huevos or 1))
        return jsonify({
            'success': True,
            'huevo': {
                'id': huevo.id,
                'peso': huevo.peso,
                'roto': huevo.roto,
                'categoria': (
                    huevo.categoria.nombre
                    if huevo.categoria
                    else ('No disponible' if huevo.roto else 'Sin clasificar')
                ),
                'lote_id': lote.id,
                'pesa_id': pesa.id
            },
            'stats': {
                'total_huevos': int(lote.total_huevos or 0),
                'huevos_rotos': int(lote.huevos_rotos or 0),
                'peso_total': float(lote.total_peso or 0),
                'peso_promedio': promedio
            }
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500

@inventario_bp.route('/lotes/<int:lote_id>/estadisticas', methods=['GET'])
@login_required
def obtener_estadisticas(lote_id):
    """Obtener estadisticas actualizadas del lote sin recargar la pagina"""
    LoteRecoleccion.query.get_or_404(lote_id)

    total_huevos = (
        db.session.query(db.func.count(Huevo.id))
        .filter(Huevo.lote_id == lote_id, Huevo.roto == False)
        .scalar()
        or 0
    )
    huevos_rotos = (
        db.session.query(db.func.count(Huevo.id))
        .filter(Huevo.lote_id == lote_id, Huevo.roto == True)
        .scalar()
        or 0
    )
    peso_total = float(
        db.session.query(db.func.coalesce(db.func.sum(Huevo.peso), 0.0))
        .filter(Huevo.lote_id == lote_id, Huevo.roto == False)
        .scalar()
        or 0.0
    )
    peso_promedio = peso_total / total_huevos if total_huevos > 0 else 0.0

    return jsonify({
        'total_huevos': int(total_huevos),
        'huevos_rotos': int(huevos_rotos),
        'peso_total': float(peso_total),
        'peso_promedio': float(peso_promedio)
    })


@inventario_bp.route('/lotes/<int:lote_id>/ultimo-huevo', methods=['GET'])
@login_required
def obtener_ultimo_huevo(lote_id):
    """Obtener el ultimo huevo agregado en el lote"""
    lote = LoteRecoleccion.query.get_or_404(lote_id)

    if lote.usuario_id != current_user.id and not current_user.is_admin:
        return jsonify({'success': False, 'message': 'Sin permisos'}), 403

    after_id = request.args.get('after_id', 0, type=int)
    ultima_fila = (
        db.session.query(Huevo.id)
        .filter(Huevo.lote_id == lote_id)
        .order_by(Huevo.timestamp.desc(), Huevo.id.desc())
        .first()
    )
    ultimo_id = int(ultima_fila[0]) if ultima_fila else None
    if not ultimo_id:
        return jsonify({'success': True, 'has_new': False, 'huevo': None})

    if after_id and ultimo_id <= after_id:
        return jsonify({'success': True, 'has_new': False})

    huevo = Huevo.query.get(ultimo_id)
    promedio = float((lote.total_peso or 0) / (lote.total_huevos or 1))
    return jsonify({
        'success': True,
        'has_new': True,
        'huevo': {
            'id': huevo.id,
            'peso': huevo.peso,
            'roto': huevo.roto,
            'categoria': (
                huevo.categoria.nombre
                if huevo.categoria
                else ('No disponible' if huevo.roto else 'Sin clasificar')
            ),
            'timestamp': huevo.timestamp.isoformat()
        },
        'stats': {
            'total_huevos': int(lote.total_huevos or 0),
            'huevos_rotos': int(lote.huevos_rotos or 0),
            'peso_total': float(lote.total_peso or 0),
            'peso_promedio': promedio
        }
    })

@inventario_bp.route('/lotes/<int:lote_id>/pesar/archivo', methods=['POST'])
@login_required
def procesar_archivo_pesos(lote_id):
    """Procesar archivo .txt con pesos o .xlsx con cantidades por categoria"""
    lote = LoteRecoleccion.query.get_or_404(lote_id)

    if lote.usuario_id != current_user.id and not current_user.is_admin:
        flash('Sin permisos para modificar este lote', 'error')
        return redirect(url_for('inventario.pesar_huevos', lote_id=lote_id))

    if not lote.pesa_id:
        flash('Debes asignar una pesa al lote antes de cargar pesos', 'error')
        return redirect(url_for('inventario.pesar_huevos', lote_id=lote_id))

    try:
        contenido_archivo = request.form.get('archivo_contenido', '').strip()
        archivo_txt = request.files.get('archivo_txt')
        archivo_xlsx = request.files.get('archivo_xlsx')

        huevos_procesados = 0
        huevos_error = 0

        # Prioridad 1: archivo .xlsx con cantidades por categoria
        if archivo_xlsx and archivo_xlsx.filename:
            try:
                wb = load_workbook(archivo_xlsx, data_only=True)
                ws = wb.active
            except Exception as e:
                flash(f'No se pudo leer el archivo .xlsx: {str(e)}', 'error')
                return redirect(url_for('inventario.pesar_huevos', lote_id=lote_id))

            categorias_activas = CategoriaHuevo.query.filter_by(activo=True).all()
            categorias_por_nombre = {
                c.nombre.strip().upper(): c for c in categorias_activas
            }

            filas_invalidas = 0

            # Esperado: columna A=Categoria, columna B=Cantidad
            for row in ws.iter_rows(min_row=2, values_only=True):
                categoria_raw = row[0] if len(row) > 0 else None
                cantidad_raw = row[1] if len(row) > 1 else None

                if categoria_raw in (None, '') and cantidad_raw in (None, ''):
                    continue

                nombre_categoria = str(categoria_raw or '').strip().upper()
                if not nombre_categoria:
                    filas_invalidas += 1
                    continue

                categoria = categorias_por_nombre.get(nombre_categoria)
                if not categoria:
                    filas_invalidas += 1
                    continue

                try:
                    cantidad = int(float(cantidad_raw))
                except (TypeError, ValueError):
                    filas_invalidas += 1
                    continue

                if cantidad <= 0:
                    filas_invalidas += 1
                    continue

                # Peso estimado para mantener consistencia de estadisticas del lote
                peso_estimado = float((categoria.peso_min + categoria.peso_max) / 2.0)
                for _ in range(cantidad):
                    huevo = Huevo(
                        peso=peso_estimado,
                        lote_id=lote_id,
                        categoria_id=categoria.id,
                        roto=False
                    )
                    db.session.add(huevo)
                huevos_procesados += cantidad

            if huevos_procesados == 0 and filas_invalidas == 0:
                flash('El archivo .xlsx no contiene filas para procesar', 'error')
                return redirect(url_for('inventario.pesar_huevos', lote_id=lote_id))

            huevos_error = filas_invalidas
            db.session.commit()
            lote.actualizar_estadisticas()
            db.session.commit()

            flash(
                f'Carga masiva completada: {huevos_procesados} huevos agregados por categoria. '
                f'{huevos_error} filas invalidas.',
                'success'
            )
            return redirect(url_for('inventario.pesar_huevos', lote_id=lote_id))

        # Prioridad: archivo .txt subido
        if archivo_txt and archivo_txt.filename:
            try:
                contenido_bytes = archivo_txt.read()
                contenido_archivo = contenido_bytes.decode('utf-8-sig')
            except UnicodeDecodeError:
                contenido_archivo = contenido_bytes.decode('latin-1', errors='ignore')

        if not contenido_archivo:
            flash('Debes subir un archivo .txt/.xlsx o pegar contenido para procesar', 'error')
            return redirect(url_for('inventario.pesar_huevos', lote_id=lote_id))

        for linea in contenido_archivo.splitlines():
            peso_str = linea.strip()
            if peso_str:
                try:
                    peso = float(peso_str.replace(',', '.'))
                    if peso > 0:
                        huevo = Huevo(
                            peso=peso,
                            lote_id=lote_id
                        )
                        huevo.clasificar()
                        db.session.add(huevo)
                        huevos_procesados += 1
                    else:
                        huevos_error += 1
                except ValueError:
                    huevos_error += 1

        db.session.commit()

        lote.actualizar_estadisticas()
        db.session.commit()

        flash(f'Procesados {huevos_procesados} huevos correctamente. {huevos_error} errores.', 'success')

    except Exception as e:
        db.session.rollback()
        flash(f'Error al procesar el archivo: {str(e)}', 'error')

    return redirect(url_for('inventario.pesar_huevos', lote_id=lote_id))

@inventario_bp.route('/lotes/<int:lote_id>/pesar/plantilla-categorias.xlsx')
@login_required
def descargar_plantilla_categorias(lote_id):
    """Descargar plantilla Excel para carga masiva por categoria"""
    lote = LoteRecoleccion.query.get_or_404(lote_id)

    if lote.usuario_id != current_user.id and not current_user.is_admin:
        flash('Sin permisos para acceder a este lote', 'error')
        return redirect(url_for('inventario.lotes'))

    categorias = CategoriaHuevo.query.filter_by(activo=True).order_by(CategoriaHuevo.peso_min.asc()).all()
    rows = [
        [c.nombre, 0, float(c.peso_min), float(c.peso_max)]
        for c in categorias
    ]

    return create_excel_response(
        f'plantilla_carga_masiva_{lote.numero_lote}.xlsx',
        'Plantilla',
        ['Categoria', 'Cantidad', 'Peso Min (referencia)', 'Peso Max (referencia)'],
        rows
    )

@inventario_bp.route('/huevos/<int:huevo_id>/marcar-roto', methods=['POST'])
@login_required
def marcar_huevo_roto(huevo_id):
    """Marcar un huevo como roto"""
    try:
        huevo = Huevo.query.get_or_404(huevo_id)
        
        # Verificar permisos
        if huevo.lote.usuario_id != current_user.id and not current_user.is_admin:
            return jsonify({'success': False, 'message': 'Sin permisos'})
        
        huevo.roto = True
        huevo.categoria_id = None  # Quitar categorÃ­a si estÃ¡ roto
        
        db.session.commit()
        
        # Actualizar estadÃ­sticas del lote
        huevo.lote.actualizar_estadisticas()
        db.session.commit()
        
        return jsonify({'success': True, 'message': 'Huevo marcado como roto'})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)})


@inventario_bp.route('/huevos/<int:huevo_id>/editar', methods=['POST'])
@login_required
def editar_huevo(huevo_id):
    """Editar un huevo en un lote de pesaje."""
    try:
        huevo = Huevo.query.get_or_404(huevo_id)
        lote = huevo.lote

        if lote.usuario_id != current_user.id and not current_user.is_admin:
            return jsonify({'success': False, 'message': 'Sin permisos'})

        if lote.estado != 'EN_PROCESO':
            return jsonify({'success': False, 'message': 'Solo se pueden editar huevos de lotes en proceso'})

        if huevo.vendido:
            return jsonify({'success': False, 'message': 'No se puede editar un huevo ya vendido'})

        payload = request.get_json(silent=True) or {}
        peso = float(payload.get('peso', huevo.peso or 0))
        roto = bool(payload.get('roto', huevo.roto))

        if not roto and peso <= 0:
            return jsonify({'success': False, 'message': 'Peso invalido'})

        if roto and peso <= 0:
            peso = 0

        huevo.peso = peso
        huevo.roto = roto

        if huevo.roto:
            huevo.categoria_id = None
        else:
            huevo.clasificar()

        db.session.commit()

        lote.actualizar_estadisticas()
        db.session.commit()

        return jsonify({
            'success': True,
            'message': 'Huevo actualizado correctamente',
            'huevo': {
                'id': huevo.id,
                'peso': huevo.peso,
                'roto': huevo.roto,
                'categoria': huevo.categoria.nombre if huevo.categoria else 'Sin clasificar'
            }
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)})


@inventario_bp.route('/huevos/<int:huevo_id>/eliminar', methods=['POST'])
@login_required
def eliminar_huevo(huevo_id):
    """Eliminar un huevo registrado en un lote."""
    try:
        huevo = Huevo.query.get_or_404(huevo_id)
        lote = huevo.lote

        if lote.usuario_id != current_user.id and not current_user.is_admin:
            return jsonify({'success': False, 'message': 'Sin permisos'})

        if lote.estado != 'EN_PROCESO':
            return jsonify({'success': False, 'message': 'Solo se pueden eliminar huevos de lotes en proceso'})

        if huevo.vendido:
            return jsonify({'success': False, 'message': 'No se puede eliminar un huevo ya vendido'})

        db.session.delete(huevo)
        db.session.commit()

        lote.actualizar_estadisticas()
        db.session.commit()

        return jsonify({'success': True, 'message': 'Huevo eliminado correctamente'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)})

@inventario_bp.route('/lotes/<int:lote_id>/ver')
@login_required
def ver_detalle_lote(lote_id):
    """Ver detalles completos de un lote"""
    lote = LoteRecoleccion.query.get_or_404(lote_id)
    
    # EstadÃ­sticas del lote
    huevos = Huevo.query.filter_by(lote_id=lote_id).all()
    stats_lote = {
        'total_huevos': len([h for h in huevos if not h.roto]),
        'huevos_rotos': len([h for h in huevos if h.roto]),
        'peso_total': sum([h.peso for h in huevos if not h.roto]),
        'peso_promedio': sum([h.peso for h in huevos if not h.roto]) / max(len([h for h in huevos if not h.roto]), 1)
    }
    
    # DistribuciÃ³n por categorÃ­as
    distribucion = db.session.query(
        CategoriaHuevo.nombre,
        CategoriaHuevo.peso_min,
        CategoriaHuevo.peso_max,
        db.func.count(Huevo.id).label('cantidad'),
        db.func.sum(Huevo.peso).label('peso_total')
    ).join(Huevo).filter(
        Huevo.lote_id == lote_id,
        Huevo.roto == False
    ).group_by(CategoriaHuevo.id, CategoriaHuevo.nombre).all()
    
    return render_template('inventario/detalle_lote.html',
                         lote=lote,
                         stats_lote=stats_lote,
                         distribucion=distribucion,
                         huevos=huevos)

@inventario_bp.route('/lotes/<int:lote_id>/completar', methods=['POST'])
@login_required
def completar_lote(lote_id):
    """Completar un lote de recolecciÃ³n"""
    try:
        lote = LoteRecoleccion.query.get_or_404(lote_id)
        
        if lote.usuario_id != current_user.id and not current_user.is_admin:
            flash('Sin permisos para completar este lote', 'error')
            return redirect(url_for('inventario.pesar_huevos', lote_id=lote_id))
        
        lote.completar_lote()
        db.session.commit()
        
        # Actualizar inventario general
        InventarioHuevos.actualizar_inventario()
        
        flash(f'Lote {lote.numero_lote} completado correctamente', 'success')
        return redirect(url_for('inventario.lotes'))
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error al completar el lote: {str(e)}', 'error')
        return redirect(url_for('inventario.pesar_huevos', lote_id=lote_id))
















